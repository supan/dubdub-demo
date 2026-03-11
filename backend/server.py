from fastapi import FastAPI, APIRouter, HTTPException, Depends, Header, Response, Request, UploadFile, File, Form, Query
from fastapi.responses import JSONResponse
from dotenv import load_dotenv
from starlette.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
import os
import logging
import asyncio
from pathlib import Path
from pydantic import BaseModel, Field
from typing import List, Optional, Dict, Any, Set, Tuple
import uuid
from datetime import datetime, timezone, timedelta
import httpx
import re

ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

# MongoDB connection
mongo_url = os.environ['MONGO_URL']
client = AsyncIOMotorClient(mongo_url)
db = client[os.environ['DB_NAME']]

# Create the main app without a prefix
app = FastAPI()

# Health check endpoint for Kubernetes (must be at root, not /api)
@app.get("/health")
async def health_check():
    """Health check endpoint for Kubernetes liveness/readiness probes"""
    return {"status": "healthy"}

# Create a router with the /api prefix
api_router = APIRouter(prefix="/api")

# ==================== VERSION COMPARISON HELPER ====================

def parse_version(version_str: str) -> Tuple[int, int, int]:
    """Parse version string like '1.3.7' into tuple (1, 3, 7)"""
    # Remove 'v' prefix if present
    version_str = version_str.lstrip('v').strip()
    
    # Match semantic version pattern
    match = re.match(r'^(\d+)\.(\d+)\.(\d+)', version_str)
    if match:
        return (int(match.group(1)), int(match.group(2)), int(match.group(3)))
    
    # Fallback: try to parse whatever we can
    parts = version_str.split('.')
    try:
        major = int(parts[0]) if len(parts) > 0 else 0
        minor = int(parts[1]) if len(parts) > 1 else 0
        patch = int(parts[2]) if len(parts) > 2 else 0
        return (major, minor, patch)
    except (ValueError, IndexError):
        return (0, 0, 0)

def compare_versions(version1: str, version2: str) -> int:
    """
    Compare two version strings.
    Returns:
        -1 if version1 < version2
         0 if version1 == version2
         1 if version1 > version2
    """
    v1 = parse_version(version1)
    v2 = parse_version(version2)
    
    if v1 < v2:
        return -1
    elif v1 > v2:
        return 1
    return 0

def is_version_compatible(app_version: str, min_required_version: str) -> bool:
    """Check if app_version meets the minimum required version"""
    return compare_versions(app_version, min_required_version) >= 0

# Minimum app version required for each playable TYPE
# When introducing breaking changes to a format, update this mapping
TYPE_MIN_VERSION = {
    "text": "1.0.0",
    "image_text": "1.0.0",
    "video_text": "1.0.0",
    "guess_the_x": "1.0.0",
    "this_or_that": "1.0.0",
    "chess_mate_in_2": "1.3.8",  # New Lichess format with opponent-first moves
}

def get_type_min_version(playable_type: str) -> str:
    """Get minimum app version required for a playable type"""
    return TYPE_MIN_VERSION.get(playable_type, "1.0.0")

# ==================== ASYNC TASK MANAGER ====================

class TaskManager:
    """
    Manages background tasks with tracking and graceful shutdown.
    
    Features:
    - Tracks all pending tasks
    - Ensures tasks complete on shutdown
    - Retries failed tasks (configurable)
    - Logs task status for monitoring
    """
    
    def __init__(self, max_retries: int = 2, retry_delay: float = 0.5):
        self._pending_tasks: Set[asyncio.Task] = set()
        self._max_retries = max_retries
        self._retry_delay = retry_delay
        self._shutdown = False
        self._lock = asyncio.Lock()
    
    async def create_task(self, coro, task_name: str = "unnamed"):
        """Create and track a background task with retry support"""
        if self._shutdown:
            logging.warning(f"Task '{task_name}' rejected - shutdown in progress")
            return None
        
        async def wrapped_task():
            retries = 0
            while retries <= self._max_retries:
                try:
                    await coro
                    logging.debug(f"Task '{task_name}' completed successfully")
                    return
                except Exception as e:
                    retries += 1
                    if retries <= self._max_retries:
                        logging.warning(f"Task '{task_name}' failed (attempt {retries}/{self._max_retries + 1}): {e}")
                        await asyncio.sleep(self._retry_delay * retries)  # Exponential backoff
                    else:
                        logging.error(f"Task '{task_name}' failed permanently after {self._max_retries + 1} attempts: {e}")
        
        task = asyncio.create_task(wrapped_task())
        
        async with self._lock:
            self._pending_tasks.add(task)
        
        # Auto-cleanup when task completes
        task.add_done_callback(lambda t: asyncio.create_task(self._remove_task(t)))
        
        return task
    
    async def _remove_task(self, task: asyncio.Task):
        """Remove completed task from tracking set"""
        async with self._lock:
            self._pending_tasks.discard(task)
    
    @property
    def pending_count(self) -> int:
        """Number of tasks still pending"""
        return len(self._pending_tasks)
    
    async def shutdown(self, timeout: float = 10.0):
        """
        Graceful shutdown - wait for all pending tasks to complete.
        
        Args:
            timeout: Maximum seconds to wait before force-cancelling tasks
        """
        self._shutdown = True
        
        if not self._pending_tasks:
            logging.info("TaskManager: No pending tasks, shutdown complete")
            return
        
        logging.info(f"TaskManager: Waiting for {len(self._pending_tasks)} pending tasks...")
        
        try:
            # Wait for all tasks with timeout
            async with self._lock:
                tasks = list(self._pending_tasks)
            
            done, pending = await asyncio.wait(
                tasks,
                timeout=timeout,
                return_when=asyncio.ALL_COMPLETED
            )
            
            if pending:
                logging.warning(f"TaskManager: {len(pending)} tasks didn't complete in time, cancelling...")
                for task in pending:
                    task.cancel()
                # Wait briefly for cancellation
                await asyncio.gather(*pending, return_exceptions=True)
            
            logging.info(f"TaskManager: Shutdown complete. {len(done)} tasks finished, {len(pending)} cancelled")
            
        except Exception as e:
            logging.error(f"TaskManager: Error during shutdown: {e}")


# Global task manager instance
task_manager = TaskManager(max_retries=2, retry_delay=0.5)


# ==================== APP LIFECYCLE EVENTS ====================

@app.on_event("startup")
async def startup_event():
    """Initialize app on startup"""
    logging.info("Application starting up...")
    # Create indexes
    try:
        await db.users.create_index("user_id", unique=True)
        await db.users.create_index("email", unique=True)
        await db.playables.create_index("playable_id", unique=True)
        await db.playables.create_index("category")
        await db.playables.create_index("type")  # Index for type filtering
        await db.playables.create_index("weight")
        await db.playables.create_index("created_at")  # Index for sorting by date
        await db.playables.create_index([("category", 1), ("type", 1)])  # Compound index for filtered queries
        await db.user_progress.create_index([("user_id", 1), ("playable_id", 1)], unique=True)
        await db.sessions.create_index("session_token", unique=True)
        await db.sessions.create_index("expires_at", expireAfterSeconds=0)
        logging.info("Database indexes created successfully")
    except Exception as e:
        logging.error(f"Error creating indexes: {e}")


@app.on_event("shutdown")
async def shutdown_event():
    """Graceful shutdown - ensure all background tasks complete"""
    logging.info("Application shutting down...")
    await task_manager.shutdown(timeout=10.0)
    logging.info("Shutdown complete")


# ==================== MODELS ====================

class User(BaseModel):
    user_id: str
    email: Optional[str] = None  # Can be None for Apple users who hide their email
    name: str
    picture: Optional[str] = None
    total_played: int = 0
    correct_answers: int = 0
    current_streak: int = 0
    best_streak: int = 0
    selected_categories: Optional[List[str]] = None  # User's selected categories
    onboarding_complete: bool = False  # Whether user has completed category selection
    created_at: Optional[datetime] = None  # Made optional for backwards compatibility

class SessionDataResponse(BaseModel):
    id: str
    email: str
    name: str
    picture: str
    session_token: str

from enum import Enum

class PlayableStatus(str, Enum):
    ACTIVE = "active"
    INACTIVE = "inactive"

# Define valid answer_types for each question type
# None means the answer_type is auto-determined and not user-selectable
PLAYABLE_TYPE_CONFIG = {
    "text": {
        "valid_answer_types": ["mcq", "text_input"],
        "default_answer_type": "mcq",
        "description": "Text-only question"
    },
    "image": {
        "valid_answer_types": ["mcq", "text_input"],
        "default_answer_type": "mcq",
        "description": "Image-only question"
    },
    "image_text": {
        "valid_answer_types": ["mcq", "text_input"],
        "default_answer_type": "mcq",
        "description": "Image with text question"
    },
    "video": {
        "valid_answer_types": ["mcq", "text_input"],
        "default_answer_type": "mcq",
        "description": "Video-only question"
    },
    "video_text": {
        "valid_answer_types": ["mcq", "text_input"],
        "default_answer_type": "mcq",
        "description": "Video with text question"
    },
    "guess_the_x": {
        "valid_answer_types": ["progressive_reveal"],
        "default_answer_type": "progressive_reveal",
        "description": "Progressive hint reveal game - user guesses after each hint"
    },
    "chess_mate_in_2": {
        "valid_answer_types": ["chess_moves"],
        "default_answer_type": "chess_moves",
        "description": "Chess puzzle - user makes moves on board"
    },
    "this_or_that": {
        "valid_answer_types": ["tap_select"],
        "default_answer_type": "tap_select",
        "description": "Choose between two options by tapping"
    },
    "wordle": {
        "valid_answer_types": ["wordle_grid"],
        "default_answer_type": "wordle_grid",
        "description": "Wordle-style word guessing game"
    }
}

def get_valid_answer_type(playable_type: str, requested_answer_type: Optional[str] = None) -> str:
    """Get valid answer_type for a playable type, validating or defaulting as needed."""
    config = PLAYABLE_TYPE_CONFIG.get(playable_type)
    if not config:
        raise ValueError(f"Unknown playable type: {playable_type}")
    
    valid_types = config["valid_answer_types"]
    default_type = config["default_answer_type"]
    
    # If only one valid type, always use it (auto-determined types)
    if len(valid_types) == 1:
        return valid_types[0]
    
    # If requested type is valid, use it
    if requested_answer_type and requested_answer_type in valid_types:
        return requested_answer_type
    
    # Otherwise use default
    return default_type

class Playable(BaseModel):
    playable_id: str
    type: str  # "text", "image_text", "video_text", "guess_the_x", "chess_mate_in_2", "this_or_that"
    answer_type: str  # "mcq", "text_input"
    category: str
    title: str
    question: Dict[str, Any]  # {text?, video_url?, image_base64?, image_url?}
    options: Optional[List[str]] = None  # For MCQ
    correct_answer: str
    alternate_answers: Optional[List[str]] = None  # For text_input: spelling variants, short forms
    answer_explanation: Optional[str] = None  # Brief explanation of the answer
    hints: Optional[List[str]] = None  # For guess_the_x: 3-5 hints revealed progressively
    difficulty: str = "medium"
    weight: int = 0  # Ranking weight: higher = shown first (0 = default/lowest priority)
    status: PlayableStatus = PlayableStatus.ACTIVE  # Status: active or inactive
    created_at: datetime

class AnswerSubmission(BaseModel):
    answer: str
    time_taken: Optional[float] = None  # Time taken to answer in seconds

class GuessAnswerSubmission(BaseModel):
    answer: str
    hint_number: int  # Which hint the user is on (1-based)
    time_taken: Optional[float] = None  # Time taken to answer in seconds

class UserProgress(BaseModel):
    user_id: str
    playable_id: str
    answered: bool
    correct: bool
    timestamp: datetime

class Category(BaseModel):
    category_id: str
    name: str
    icon: str  # Icon name (e.g., 'flask', 'globe', 'book')
    color: str  # Hex color for the category
    playable_count: int = 0

class CategorySelectionRequest(BaseModel):
    categories: List[str]  # List of category names

# ==================== AUTH HELPERS ====================

async def get_current_user_from_token(authorization: Optional[str] = Header(None)) -> Optional[User]:
    """Get user from Authorization header token"""
    if not authorization:
        return None
    
    try:
        # Remove "Bearer " prefix if present
        session_token = authorization.replace("Bearer ", "")
        
        # Find session
        session = await db.user_sessions.find_one(
            {"session_token": session_token},
            {"_id": 0}
        )
        
        if not session:
            return None
        
        # Check if session is expired
        expires_at = session["expires_at"]
        if expires_at.tzinfo is None:
            expires_at = expires_at.replace(tzinfo=timezone.utc)
        
        if expires_at < datetime.now(timezone.utc):
            return None
        
        # Get user
        user_doc = await db.users.find_one(
            {"user_id": session["user_id"]},
            {"_id": 0}
        )
        
        if user_doc:
            return User(**user_doc)
        
        return None
    except Exception as e:
        logging.error(f"Error getting current user: {e}")
        return None

async def require_auth(authorization: Optional[str] = Header(None)) -> User:
    """Require authentication"""
    user = await get_current_user_from_token(authorization)
    if not user:
        raise HTTPException(status_code=401, detail="Not authenticated")
    return user

# ==================== AUTH ENDPOINTS ====================

@api_router.post("/auth/dev-login")
async def dev_login():
    """Development login - bypasses OAuth for testing"""
    try:
        # Hardcoded dev user
        dev_email = "supanshah51191@gmail.com"
        
        # Find or create user
        user_doc = await db.users.find_one({"email": dev_email}, {"_id": 0})
        
        if not user_doc:
            # Create dev user
            user_id = f"user_{uuid.uuid4().hex[:12]}"
            new_user = {
                "user_id": user_id,
                "email": dev_email,
                "name": "Dev User",
                "picture": None,
                "total_played": 0,
                "correct_answers": 0,
                "current_streak": 0,
                "best_streak": 0,
                "selected_categories": None,
                "onboarding_complete": False,
                "created_at": datetime.now(timezone.utc)
            }
            await db.users.insert_one(new_user)
            user_id_to_use = user_id
        else:
            user_id_to_use = user_doc["user_id"]
        
        # Create session
        dev_session_token = f"dev_session_{uuid.uuid4().hex}"
        session_doc = {
            "user_id": user_id_to_use,
            "session_token": dev_session_token,
            "expires_at": datetime.now(timezone.utc) + timedelta(days=180),
            "created_at": datetime.now(timezone.utc)
        }
        await db.user_sessions.insert_one(session_doc)
        
        return {
            "session_token": dev_session_token,
            "user": {
                "user_id": user_id_to_use,
                "email": dev_email,
                "name": "Dev User"
            }
        }
    
    except Exception as e:
        logging.error(f"Dev login error: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@api_router.post("/auth/session")
async def exchange_session(request: Request, response: Response):
    """Exchange session_id for session_token"""
    try:
        # Get session_id from header
        session_id = request.headers.get("X-Session-ID")
        
        if not session_id:
            raise HTTPException(status_code=400, detail="X-Session-ID header required")
        
        # Call Emergent Auth API
        async with httpx.AsyncClient() as client:
            auth_response = await client.get(
                "https://demobackend.emergentagent.com/auth/v1/env/oauth/session-data",
                headers={"X-Session-ID": session_id},
                timeout=10.0
            )
            
            if auth_response.status_code != 200:
                raise HTTPException(status_code=401, detail="Invalid session")
            
            user_data = auth_response.json()
        
        # Create SessionDataResponse
        session_data = SessionDataResponse(**user_data)
        
        # Check if user exists
        existing_user = await db.users.find_one(
            {"email": session_data.email},
            {"_id": 0}
        )
        
        if not existing_user:
            # Create new user
            user_id = f"user_{uuid.uuid4().hex[:12]}"
            new_user = {
                "user_id": user_id,
                "email": session_data.email,
                "name": session_data.name,
                "picture": session_data.picture,
                "total_played": 0,
                "correct_answers": 0,
                "current_streak": 0,
                "best_streak": 0,
                "selected_categories": None,
                "onboarding_complete": False,
                "created_at": datetime.now(timezone.utc)
            }
            await db.users.insert_one(new_user)
            user_id_to_use = user_id
        else:
            user_id_to_use = existing_user["user_id"]
        
        # Store session in database
        session_doc = {
            "user_id": user_id_to_use,
            "session_token": session_data.session_token,
            "expires_at": datetime.now(timezone.utc) + timedelta(days=180),
            "created_at": datetime.now(timezone.utc)
        }
        await db.user_sessions.insert_one(session_doc)
        
        # Return session data
        return {
            "session_token": session_data.session_token,
            "user": {
                "user_id": user_id_to_use,
                "email": session_data.email,
                "name": session_data.name,
                "picture": session_data.picture
            }
        }
    
    except httpx.TimeoutException:
        raise HTTPException(status_code=504, detail="Auth service timeout")
    except Exception as e:
        logging.error(f"Session exchange error: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@api_router.get("/auth/me")
async def get_me(current_user: User = Depends(require_auth)):
    """Get current user info"""
    return current_user

@api_router.post("/auth/logout")
async def logout(authorization: Optional[str] = Header(None)):
    """Logout user"""
    if not authorization:
        raise HTTPException(status_code=400, detail="No authorization header")
    
    session_token = authorization.replace("Bearer ", "")
    
    # Delete session
    await db.user_sessions.delete_one({"session_token": session_token})
    
    return {"message": "Logged out successfully"}

class AppleAuthRequest(BaseModel):
    identity_token: str
    apple_user_id: str
    email: Optional[str] = None
    name: Optional[str] = None

@api_router.post("/auth/apple")
async def apple_auth(request: AppleAuthRequest):
    """
    Handle Apple Sign In authentication.
    Verifies the identity token and creates/retrieves user account.
    Note: Apple only provides email on FIRST sign-in. Users can also hide their email.
    """
    try:
        import jwt
        from jwt import PyJWKClient
        
        # Apple's public keys URL for token verification
        APPLE_KEYS_URL = "https://appleid.apple.com/auth/keys"
        
        # Verify the identity token
        try:
            # Get Apple's public keys
            jwks_client = PyJWKClient(APPLE_KEYS_URL)
            signing_key = jwks_client.get_signing_key_from_jwt(request.identity_token)
            
            # Decode and verify the token
            decoded = jwt.decode(
                request.identity_token,
                signing_key.key,
                algorithms=["RS256"],
                audience="com.emergent.invin",  # Your bundle ID
                issuer="https://appleid.apple.com"
            )
            
            # Extract user info from token
            apple_user_id = decoded.get("sub")
            # Apple may not provide email on subsequent sign-ins or if user chose "Hide My Email"
            email = decoded.get("email") or request.email
            
            # If no email provided, generate a placeholder based on apple_user_id
            # This ensures we can still create a valid user record
            if not email:
                # Use a recognizable pattern that won't conflict with real emails
                email = f"apple_{apple_user_id[:12]}@privaterelay.apple.local"
                logging.info(f"Apple auth: No email provided, using generated email: {email}")
            
        except jwt.ExpiredSignatureError:
            raise HTTPException(status_code=401, detail="Apple token has expired")
        except jwt.InvalidTokenError as e:
            logging.error(f"Apple token validation error: {e}")
            raise HTTPException(status_code=401, detail="Invalid Apple token")
        
        # Check if user exists by apple_user_id first (most reliable), then by email
        existing_user = await db.users.find_one(
            {"$or": [
                {"apple_user_id": apple_user_id},
                {"email": email}
            ]},
            {"_id": 0}
        )
        
        if not existing_user:
            # Create new user
            user_id = f"user_{uuid.uuid4().hex[:12]}"
            new_user = {
                "user_id": user_id,
                "apple_user_id": apple_user_id,
                "email": email,
                "name": request.name or "Apple User",
                "picture": None,
                "total_played": 0,
                "correct_answers": 0,
                "current_streak": 0,
                "best_streak": 0,
                "selected_categories": None,
                "onboarding_complete": False,
                "created_at": datetime.now(timezone.utc)
            }
            await db.users.insert_one(new_user)
            user_id_to_use = user_id
            user_email = email
            user_name = request.name or "Apple User"
        else:
            user_id_to_use = existing_user["user_id"]
            user_email = existing_user.get("email") or email
            user_name = existing_user.get("name") or request.name or "Apple User"
            
            # Build update fields
            update_fields = {}
            
            # Update apple_user_id if not set (user previously logged in with Google)
            if not existing_user.get("apple_user_id"):
                update_fields["apple_user_id"] = apple_user_id
            
            # Update email if existing email is null/empty or is a generated placeholder
            existing_email = existing_user.get("email")
            if not existing_email or existing_email.endswith("@privaterelay.apple.local"):
                if email and not email.endswith("@privaterelay.apple.local"):
                    update_fields["email"] = email
                    user_email = email
            
            # Update name if provided and existing is default
            if request.name and existing_user.get("name") == "Apple User":
                update_fields["name"] = request.name
                user_name = request.name
            
            # Apply updates if any
            if update_fields:
                await db.users.update_one(
                    {"user_id": user_id_to_use},
                    {"$set": update_fields}
                )
                logging.info(f"Apple auth: Updated user {user_id_to_use} with fields: {list(update_fields.keys())}")
        
        # Create session token
        session_token = f"apple_{uuid.uuid4().hex}"
        
        # Store session in database
        session_doc = {
            "user_id": user_id_to_use,
            "session_token": session_token,
            "expires_at": datetime.now(timezone.utc) + timedelta(days=180),
            "created_at": datetime.now(timezone.utc),
            "auth_provider": "apple"
        }
        await db.user_sessions.insert_one(session_doc)
        
        # Return session data
        return {
            "session_token": session_token,
            "user": {
                "user_id": user_id_to_use,
                "email": user_email,
                "name": user_name,
                "picture": None
            }
        }
        
    except HTTPException:
        raise
    except Exception as e:
        logging.error(f"Apple auth error: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@api_router.delete("/auth/delete-account")
async def delete_account(current_user: User = Depends(require_auth)):
    """
    Delete user account and all associated data.
    This permanently removes:
    - User profile from users collection
    - All progress records from user_progress collection
    - All active sessions from user_sessions collection
    """
    try:
        user_id = current_user.user_id
        
        # Delete user progress records
        progress_result = await db.user_progress.delete_many({"user_id": user_id})
        logging.info(f"Deleted {progress_result.deleted_count} progress records for user {user_id}")
        
        # Delete all user sessions
        sessions_result = await db.user_sessions.delete_many({"user_id": user_id})
        logging.info(f"Deleted {sessions_result.deleted_count} sessions for user {user_id}")
        
        # Delete the user account
        user_result = await db.users.delete_one({"user_id": user_id})
        logging.info(f"Deleted user account: {user_id}")
        
        if user_result.deleted_count == 0:
            raise HTTPException(status_code=404, detail="User not found")
        
        return {
            "message": "Account deleted successfully",
            "deleted": {
                "user": True,
                "progress_records": progress_result.deleted_count,
                "sessions": sessions_result.deleted_count
            }
        }
        
    except HTTPException:
        raise
    except Exception as e:
        logging.error(f"Error deleting account for user {current_user.user_id}: {e}")
        raise HTTPException(status_code=500, detail="Failed to delete account")

# ==================== CATEGORY ENDPOINTS ====================

# Valid Ionicons names (comprehensive list for validation)
# Source: https://ionic.io/ionicons
VALID_IONICONS = {
    # General
    "add", "add-circle", "remove", "remove-circle", "close", "close-circle",
    "checkmark", "checkmark-circle", "checkmark-done", "alert", "alert-circle",
    "information", "information-circle", "help", "help-circle", "warning",
    
    # Navigation & Actions
    "arrow-back", "arrow-forward", "arrow-up", "arrow-down", "chevron-back",
    "chevron-forward", "chevron-up", "chevron-down", "caret-back", "caret-forward",
    "caret-up", "caret-down", "menu", "ellipsis-horizontal", "ellipsis-vertical",
    "options", "settings", "cog", "build", "construct", "hammer", "key",
    
    # Media & Files
    "play", "pause", "stop", "volume-high", "volume-low", "volume-mute",
    "mic", "mic-off", "camera", "videocam", "image", "images", "film",
    "musical-notes", "musical-note", "headset", "radio", "tv", "desktop",
    "laptop", "phone-portrait", "tablet-portrait", "watch", "print",
    "document", "documents", "folder", "folder-open", "file-tray",
    "archive", "trash", "download", "cloud-download", "upload", "cloud-upload",
    "share", "share-social", "link", "copy", "clipboard", "cut",
    
    # Communication
    "mail", "mail-open", "send", "chatbox", "chatbubble", "chatbubbles",
    "call", "notifications", "notifications-off", "megaphone", "at",
    
    # People & Social
    "person", "person-add", "people", "body", "man", "woman",
    "happy", "sad", "skull", "heart", "heart-dislike", "thumbs-up", "thumbs-down",
    
    # Nature & Weather
    "sunny", "moon", "cloudy", "rainy", "thunderstorm", "snow", "flame", "bonfire",
    "water", "leaf", "flower", "rose", "earth", "globe", "planet", "star",
    
    # Objects
    "home", "business", "storefront", "cafe", "restaurant", "fast-food", "pizza", "beer",
    "wine", "ice-cream", "nutrition", "fitness", "medkit", "bandage", "pulse",
    "eye", "eye-off", "glasses", "shirt", "gift", "pricetag", "pricetags",
    "cart", "bag", "basket", "wallet", "card", "cash", "calculator", "receipt",
    "barcode", "qr-code", "scan", "sparkles", "ribbon", "medal", "trophy",
    
    # Transportation
    "airplane", "car", "bus", "train", "subway", "boat", "bicycle", "walk",
    "navigate", "compass", "map", "location", "pin", "flag",
    
    # Technology & Science
    "bulb", "flashlight", "flash", "battery-charging", "battery-full",
    "wifi", "bluetooth", "cellular", "radio-button-on", "toggle",
    "hardware-chip", "server", "terminal", "code", "code-slash",
    "git-branch", "git-commit", "git-merge", "git-pull-request",
    "logo-github", "logo-javascript", "logo-python", "logo-react",
    "flask", "beaker", "nuclear", "magnet", "prism", "cube", "shapes",
    
    # Sports & Games
    "football", "american-football", "basketball", "baseball", "tennisball",
    "golf", "fish", "game-controller", "dice", "extension-puzzle",
    
    # Education & Culture
    "book", "library", "school", "easel", "color-palette", "brush", "pencil",
    "create", "newspaper", "reader", "language", "text",
    
    # Time & Calendar
    "time", "timer", "stopwatch", "hourglass", "alarm", "calendar", "today",
    
    # Security
    "lock-closed", "lock-open", "shield", "shield-checkmark", "finger-print",
    
    # Misc
    "infinite", "rocket", "telescope", "binoculars", "aperture", "contrast",
    "layers", "grid", "apps", "analytics", "bar-chart", "pie-chart", "stats-chart",
    "trending-up", "trending-down", "podium", "funnel", "filter",
    "refresh", "reload", "sync", "repeat", "shuffle", "swap-horizontal",
    "expand", "contract", "resize", "move", "crop", "color-fill",
    "attach", "paperclip", "push", "pulse", "log-in", "log-out", "exit",
    "enter", "return-down-back", "save", "search", "bug", "skull-outline",
}

def is_valid_icon(icon_name: str) -> bool:
    """Check if icon name is a valid Ionicons name"""
    if not icon_name:
        return False
    # Also accept outline/sharp variants
    base_icon = icon_name.replace("-outline", "").replace("-sharp", "")
    return icon_name in VALID_IONICONS or base_icon in VALID_IONICONS

# Default category icons and colors mapping (used for initialization)
DEFAULT_CATEGORY_STYLES = {
    "SCIENCE": {"icon": "flask", "color": "#4CAF50"},
    "GEOGRAPHY": {"icon": "globe", "color": "#2196F3"},
    "HISTORY": {"icon": "time", "color": "#FF9800"},
    "LITERATURE": {"icon": "book", "color": "#9C27B0"},
    "SPORTS": {"icon": "american-football", "color": "#F44336"},
    "MUSIC": {"icon": "musical-notes", "color": "#E91E63"},
    "ART": {"icon": "color-palette", "color": "#00BCD4"},
    "MOVIES": {"icon": "film", "color": "#795548"},
    "TECHNOLOGY": {"icon": "hardware-chip", "color": "#607D8B"},
    "FOOD": {"icon": "restaurant", "color": "#FF5722"},
    "NATURE": {"icon": "leaf", "color": "#8BC34A"},
    "ANIMALS": {"icon": "paw", "color": "#FFEB3B"},
    "MATHEMATICS": {"icon": "calculator", "color": "#3F51B5"},
    "MATHS": {"icon": "calculator", "color": "#3F51B5"},
    "LANGUAGES": {"icon": "language", "color": "#009688"},
    "GENERAL": {"icon": "help-circle", "color": "#9E9E9E"},
    "CRICKET": {"icon": "baseball", "color": "#4CAF50"},
    "POP CULTURE": {"icon": "star", "color": "#E91E63"},
    "CHESS": {"icon": "extension-puzzle", "color": "#607D8B"},
    "BOLLYWOOD": {"icon": "videocam", "color": "#FF5722"},
    "FOOTBALL": {"icon": "football", "color": "#4CAF50"},
    "POLITICS": {"icon": "podium", "color": "#9C27B0"},
    "ENTERTAINMENT": {"icon": "tv", "color": "#E91E63"},
    "GAMING": {"icon": "game-controller", "color": "#00BCD4"},
    "TRIVIA": {"icon": "bulb", "color": "#FFEB3B"},
}

def get_default_category_style(category_name: str) -> dict:
    """Get default icon and color for a category"""
    upper_name = category_name.upper()
    if upper_name in DEFAULT_CATEGORY_STYLES:
        return DEFAULT_CATEGORY_STYLES[upper_name]
    # Default style for unknown categories
    return {"icon": "help-circle", "color": "#00FF87"}

# Model for adding a new category
class AddCategoryRequest(BaseModel):
    name: str
    icon: Optional[str] = "help-circle"
    color: Optional[str] = "#00FF87"
    description: Optional[str] = None  # Optional description for the category

@api_router.get("/categories")
async def get_categories(current_user: User = Depends(require_auth)):
    """Get all available categories from the managed categories collection"""
    try:
        # Get categories from the managed collection
        categories_cursor = db.categories.find({}, {"_id": 0}).sort("name", 1)
        categories = await categories_cursor.to_list(100)
        
        # Get playable counts for each category
        pipeline = [
            {"$group": {"_id": "$category", "count": {"$sum": 1}}}
        ]
        category_counts = await db.playables.aggregate(pipeline).to_list(100)
        count_map = {c["_id"]: c["count"] for c in category_counts}
        
        # Add playable_count to each category
        for cat in categories:
            cat["playable_count"] = count_map.get(cat["name"], 0)
        
        return {"categories": categories}
    except Exception as e:
        logging.error(f"Error fetching categories: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@api_router.get("/categories/list")
async def get_categories_list():
    """Get all category names (public endpoint for validation)"""
    try:
        categories = await db.categories.find({}, {"_id": 0, "name": 1}).to_list(100)
        return {"categories": [c["name"] for c in categories]}
    except Exception as e:
        logging.error(f"Error fetching categories list: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@api_router.post("/categories/select")
async def select_categories(
    request: CategorySelectionRequest,
    current_user: User = Depends(require_auth)
):
    """Save user's selected categories (minimum 3 required)"""
    try:
        if len(request.categories) < 3:
            raise HTTPException(
                status_code=400, 
                detail="Please select at least 3 categories"
            )
        
        # Update user with selected categories
        await db.users.update_one(
            {"user_id": current_user.user_id},
            {"$set": {
                "selected_categories": request.categories,
                "onboarding_complete": True
            }}
        )
        
        return {
            "success": True,
            "message": "Categories saved successfully",
            "selected_categories": request.categories
        }
    except HTTPException:
        raise
    except Exception as e:
        logging.error(f"Error saving categories: {e}")
        raise HTTPException(status_code=500, detail=str(e))

# ==================== PLAYABLE ENDPOINTS ====================

@api_router.get("/playables/feed")
async def get_playables_feed(
    limit: int = 10,
    app_version: Optional[str] = None,
    current_user: User = Depends(require_auth)
):
    """Get playables feed - weight-based ranking with version and category filtering
    
    Ranking Logic:
    1. Playables are sorted by 'weight' field (descending - higher weight = shown first)
    2. Playables with same weight are shown in random order
    3. Already played/skipped playables are excluded
    4. Playables whose TYPE requires a newer app version are excluded
    5. Only playables from user's selected categories are shown
    
    Weight Guidelines:
    - weight=0 (default): Normal playables, shown randomly after weighted ones
    - weight=1-10: Low priority featured content
    - weight=11-50: Medium priority featured content  
    - weight=51-100: High priority featured content
    - weight=100+: Top priority (pinned content, always shown first)
    
    Version Filtering:
    - If app_version is provided, playables are filtered by their TYPE's min version requirement
    - Version requirements are defined per TYPE in TYPE_MIN_VERSION mapping
    - Example: chess_mate_in_2 requires 1.3.8+, text requires 1.0.0+
    
    Category Filtering:
    - If user has selected categories, only show playables from those categories
    - If no categories selected (legacy users), show all playables
    """
    try:
        # Get user's played/skipped playable IDs
        played_records = await db.user_progress.find(
            {"user_id": current_user.user_id},
            {"playable_id": 1}
        ).to_list(length=10000)
        played_ids = list({r["playable_id"] for r in played_records})
        
        # Get user's selected categories
        selected_categories = current_user.selected_categories
        
        # DEBUG: Log detailed info about user and categories
        logging.info(f"Feed for user {current_user.user_id}: excluding {len(played_ids)} playables, app_version={app_version}")
        logging.info(f"User selected_categories: {selected_categories}, onboarding_complete: {current_user.onboarding_complete}")
        
        # Build base match criteria - only show ACTIVE playables
        match_criteria: Dict[str, Any] = {
            "playable_id": {"$nin": played_ids},
            "status": {"$ne": "inactive"}  # Exclude inactive, include active and legacy (no status field)
        }
        
        # Filter by selected categories (if user has selections)
        if selected_categories and len(selected_categories) > 0:
            match_criteria["category"] = {"$in": selected_categories}
            logging.info(f"Filtering to categories: {selected_categories}, match_criteria: {match_criteria}")
        else:
            logging.warning(f"User {current_user.user_id} has no selected categories - showing all playables!")
        
        # If app_version provided, exclude incompatible types
        if app_version:
            incompatible_types = [
                ptype for ptype, min_ver in TYPE_MIN_VERSION.items()
                if not is_version_compatible(app_version, min_ver)
            ]
            if incompatible_types:
                match_criteria["type"] = {"$nin": incompatible_types}
                logging.info(f"Excluding incompatible types for v{app_version}: {incompatible_types}")
        
        # Aggregation pipeline: exclude played, sort by weight desc, then randomize within same weight
        pipeline = [
            # Stage 1: Exclude already played playables, filter by categories and incompatible types
            {"$match": match_criteria},
            
            # Stage 2: Add default values for weight
            {"$addFields": {
                "weight": {"$ifNull": ["$weight", 0]}
            }},
            
            # Stage 3: Sort by weight descending (higher weight = shown first)
            {"$sort": {"weight": -1, "_id": 1}},
            
            # Stage 4: Group by weight to randomize within same weight tier
            {"$group": {
                "_id": "$weight",
                "playables": {"$push": "$$ROOT"}
            }},
            
            # Stage 5: Sort groups by weight descending
            {"$sort": {"_id": -1}},
            
            # Stage 6: Unwind back to individual playables
            {"$unwind": "$playables"},
            
            # Stage 7: Replace root with the playable document
            {"$replaceRoot": {"newRoot": "$playables"}},
            
            # Stage 8: Limit results
            {"$limit": limit}
        ]
        
        result_playables = await db.playables.aggregate(pipeline).to_list(limit)
        
        # Convert ObjectId to string
        for p in result_playables:
            p["_id"] = str(p["_id"])
        
        return result_playables
    except Exception as e:
        logging.error(f"Error fetching playables: {e}")
        raise HTTPException(status_code=500, detail=str(e))

# ==================== ASYNC DATABASE HELPERS ====================

async def save_answer_progress(user_id: str, playable_id: str, is_correct: bool):
    """Save user progress to database (fire-and-forget)"""
    try:
        progress = {
            "user_id": user_id,
            "playable_id": playable_id,
            "answered": True,
            "correct": is_correct,
            "timestamp": datetime.now(timezone.utc)
        }
        await db.user_progress.insert_one(progress)
        logging.info(f"Progress saved: user={user_id}, playable={playable_id}, correct={is_correct}")
    except Exception as e:
        logging.error(f"Failed to save progress: {e}")

async def update_user_stats(user_id: str, is_correct: bool):
    """Update user stats in database (fire-and-forget)"""
    try:
        user_doc = await db.users.find_one({"user_id": user_id}, {"_id": 0})
        if not user_doc:
            logging.error(f"User not found for stats update: {user_id}")
            return
        
        new_total_played = user_doc.get("total_played", 0) + 1
        new_correct_answers = user_doc.get("correct_answers", 0) + (1 if is_correct else 0)
        
        if is_correct:
            new_current_streak = user_doc.get("current_streak", 0) + 1
            new_best_streak = max(user_doc.get("best_streak", 0), new_current_streak)
        else:
            new_current_streak = 0
            new_best_streak = user_doc.get("best_streak", 0)
        
        await db.users.update_one(
            {"user_id": user_id},
            {"$set": {
                "total_played": new_total_played,
                "correct_answers": new_correct_answers,
                "current_streak": new_current_streak,
                "best_streak": new_best_streak
            }}
        )
        logging.info(f"Stats updated: user={user_id}, total={new_total_played}, streak={new_current_streak}")
    except Exception as e:
        logging.error(f"Failed to update user stats: {e}")

async def save_skip_progress(user_id: str, playable_id: str):
    """Save skip progress to database (fire-and-forget)"""
    try:
        progress = {
            "user_id": user_id,
            "playable_id": playable_id,
            "answered": False,
            "skipped": True,
            "correct": False,
            "timestamp": datetime.now(timezone.utc)
        }
        await db.user_progress.insert_one(progress)
        
        await db.users.update_one(
            {"user_id": user_id},
            {"$inc": {"skipped": 1}}
        )
        logging.info(f"Skip saved: user={user_id}, playable={playable_id}")
    except Exception as e:
        logging.error(f"Failed to save skip: {e}")

@api_router.post("/playables/{playable_id}/answer")
async def submit_answer(
    playable_id: str,
    submission: AnswerSubmission,
    current_user: User = Depends(require_auth)
):
    """Submit answer for a playable
    
    Flow:
    1. Validate answer (synchronous - needed for response)
    2. Save progress SYNCHRONOUSLY (critical for feed filtering)
    3. Calculate new stats (synchronous - needed for response)
    4. Update user stats asynchronously (non-critical)
    """
    try:
        # Get playable
        playable = await db.playables.find_one(
            {"playable_id": playable_id},
            {"_id": 0}
        )
        
        if not playable:
            raise HTTPException(status_code=404, detail="Playable not found")
        
        # Check answer (synchronous - needed for response)
        is_correct = submission.answer.strip().lower() == playable["correct_answer"].strip().lower()
        
        # Also check alternate answers for text input questions
        if not is_correct and playable.get("alternate_answers"):
            user_answer = submission.answer.strip().lower()
            for alt in playable["alternate_answers"]:
                if user_answer == alt.strip().lower():
                    is_correct = True
                    break
        
        # CRITICAL: Save progress SYNCHRONOUSLY to prevent duplicates in feed
        # This ensures the playable is excluded from the next feed request
        try:
            progress = {
                "user_id": current_user.user_id,
                "playable_id": playable_id,
                "answered": True,
                "correct": is_correct,
                "timestamp": datetime.now(timezone.utc)
            }
            # Add time_taken if provided
            if submission.time_taken is not None:
                progress["time_taken"] = round(submission.time_taken, 2)
            await db.user_progress.insert_one(progress)
            logging.info(f"Progress saved (sync): user={current_user.user_id}, playable={playable_id}, time={submission.time_taken}")
        except Exception as e:
            # Log but don't fail the request - duplicate key error is OK
            logging.warning(f"Progress save warning: {e}")
        
        # Calculate new stats (synchronous - needed for response)
        user_doc = await db.users.find_one({"user_id": current_user.user_id}, {"_id": 0})
        
        new_total_played = user_doc.get("total_played", 0) + 1
        new_correct_answers = user_doc.get("correct_answers", 0) + (1 if is_correct else 0)
        
        if is_correct:
            new_current_streak = user_doc.get("current_streak", 0) + 1
            new_best_streak = max(user_doc.get("best_streak", 0), new_current_streak)
        else:
            new_current_streak = 0
            new_best_streak = user_doc.get("best_streak", 0)
        
        # ASYNC: Update user stats (non-critical for feed, can be async)
        await task_manager.create_task(
            update_user_stats(current_user.user_id, is_correct),
            task_name=f"update_stats:{current_user.user_id}"
        )
        
        # Return immediately with calculated results
        return {
            "correct": is_correct,
            "correct_answer": playable["correct_answer"],
            "answer_explanation": playable.get("answer_explanation"),
            "current_streak": new_current_streak,
            "best_streak": new_best_streak,
            "total_played": new_total_played,
            "correct_answers": new_correct_answers
        }
    
    except HTTPException:
        raise
    except Exception as e:
        logging.error(f"Error submitting answer: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@api_router.post("/playables/{playable_id}/guess-answer")
async def submit_guess_answer(
    playable_id: str,
    submission: GuessAnswerSubmission,
    current_user: User = Depends(require_auth)
):
    """Submit answer for a 'Guess the X' playable - returns feedback based on hint number"""
    try:
        # Get playable
        playable = await db.playables.find_one(
            {"playable_id": playable_id},
            {"_id": 0}
        )
        
        if not playable:
            raise HTTPException(status_code=404, detail="Playable not found")
        
        if playable.get("type") != "guess_the_x":
            raise HTTPException(status_code=400, detail="This endpoint is only for 'Guess the X' playables")
        
        # Check answer
        is_correct = submission.answer.strip().lower() == playable["correct_answer"].strip().lower()
        
        # Also check alternate answers
        if not is_correct and playable.get("alternate_answers"):
            user_answer = submission.answer.strip().lower()
            for alt in playable["alternate_answers"]:
                if user_answer == alt.strip().lower():
                    is_correct = True
                    break
        
        hints = playable.get("hints", [])
        total_hints = len(hints)
        hint_number = submission.hint_number
        
        # Generate feedback message based on which hint they got it on
        feedback_messages = [
            "Incredible! You're a mind reader! 🧠",  # 1st hint
            "Impressive! You've got sharp instincts! 🎯",  # 2nd hint
            "Well done! You really know your stuff! 💪",  # 3rd hint
            "Nice work! You figured it out! 👏",  # 4th hint
            "Got it! Better late than never! ✓"  # 5th hint
        ]
        
        feedback_message = ""
        if is_correct and hint_number <= len(feedback_messages):
            feedback_message = feedback_messages[hint_number - 1]
        
        # Only save progress and update stats if correct or all hints exhausted
        if is_correct or hint_number >= total_hints:
            progress = {
                "user_id": current_user.user_id,
                "playable_id": playable_id,
                "answered": True,
                "correct": is_correct,
                "hints_used": hint_number,
                "timestamp": datetime.now(timezone.utc)
            }
            await db.user_progress.insert_one(progress)
            
            # Update user stats
            user_doc = await db.users.find_one({"user_id": current_user.user_id}, {"_id": 0})
            
            new_total_played = user_doc["total_played"] + 1
            new_correct_answers = user_doc["correct_answers"] + (1 if is_correct else 0)
            
            if is_correct:
                new_current_streak = user_doc["current_streak"] + 1
                new_best_streak = max(user_doc["best_streak"], new_current_streak)
            else:
                new_current_streak = 0
                new_best_streak = user_doc["best_streak"]
            
            await db.users.update_one(
                {"user_id": current_user.user_id},
                {"$set": {
                    "total_played": new_total_played,
                    "correct_answers": new_correct_answers,
                    "current_streak": new_current_streak,
                    "best_streak": new_best_streak
                }}
            )
            
            return {
                "correct": is_correct,
                "correct_answer": playable["correct_answer"],
                "feedback_message": feedback_message,
                "hints_used": hint_number,
                "total_hints": total_hints,
                "all_hints_exhausted": hint_number >= total_hints and not is_correct,
                "current_streak": new_current_streak,
                "best_streak": new_best_streak,
                "total_played": new_total_played,
                "correct_answers": new_correct_answers
            }
        else:
            # Wrong answer but more hints available
            return {
                "correct": False,
                "correct_answer": None,  # Don't reveal yet
                "feedback_message": "",
                "hints_used": hint_number,
                "total_hints": total_hints,
                "all_hints_exhausted": False,
                "reveal_next_hint": True
            }
    
    except HTTPException:
        raise
    except Exception as e:
        logging.error(f"Error submitting guess answer: {e}")
        raise HTTPException(status_code=500, detail=str(e))

class ChessPuzzleSubmission(BaseModel):
    solved: bool
    moves_used: int  # Number of moves used to solve

@api_router.post("/playables/{playable_id}/chess-solved")
async def submit_chess_puzzle_result(
    playable_id: str,
    submission: ChessPuzzleSubmission,
    current_user: User = Depends(require_auth)
):
    """Submit chess puzzle result"""
    try:
        # Get playable
        playable = await db.playables.find_one(
            {"playable_id": playable_id},
            {"_id": 0}
        )
        
        if not playable:
            raise HTTPException(status_code=404, detail="Playable not found")
        
        if playable.get("type") != "chess_mate_in_2":
            raise HTTPException(status_code=400, detail="This endpoint is for chess puzzles only")
        
        is_correct = submission.solved
        
        # Save progress
        progress = {
            "user_id": current_user.user_id,
            "playable_id": playable_id,
            "answered": True,
            "correct": is_correct,
            "moves_used": submission.moves_used,
            "timestamp": datetime.now(timezone.utc)
        }
        await db.user_progress.insert_one(progress)
        
        # Update user stats
        user_doc = await db.users.find_one({"user_id": current_user.user_id}, {"_id": 0})
        
        new_total_played = user_doc["total_played"] + 1
        new_correct_answers = user_doc["correct_answers"] + (1 if is_correct else 0)
        
        if is_correct:
            new_current_streak = user_doc["current_streak"] + 1
            new_best_streak = max(user_doc["best_streak"], new_current_streak)
        else:
            new_current_streak = 0
            new_best_streak = user_doc["best_streak"]
        
        await db.users.update_one(
            {"user_id": current_user.user_id},
            {"$set": {
                "total_played": new_total_played,
                "correct_answers": new_correct_answers,
                "current_streak": new_current_streak,
                "best_streak": new_best_streak
            }}
        )
        
        return {
            "correct": is_correct,
            "moves_used": submission.moves_used,
            "current_streak": new_current_streak,
            "best_streak": new_best_streak,
            "total_played": new_total_played,
            "correct_answers": new_correct_answers
        }
    
    except HTTPException:
        raise
    except Exception as e:
        logging.error(f"Error submitting chess puzzle result: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@api_router.get("/user/stats")
async def get_user_stats(current_user: User = Depends(require_auth)):
    """Get user statistics"""
    return {
        "total_played": current_user.total_played,
        "correct_answers": current_user.correct_answers,
        "current_streak": current_user.current_streak,
        "best_streak": current_user.best_streak,
        "skipped": getattr(current_user, 'skipped', 0)
    }

@api_router.post("/playables/{playable_id}/skip")
async def skip_playable(
    playable_id: str,
    current_user: User = Depends(require_auth)
):
    """Skip a playable without affecting streak
    
    Flow:
    1. Validate playable exists (synchronous)
    2. Save skip progress SYNCHRONOUSLY (critical for feed filtering)
    3. Update user stats asynchronously (non-critical)
    """
    try:
        # Get playable to verify it exists
        playable = await db.playables.find_one(
            {"playable_id": playable_id},
            {"_id": 0}
        )
        
        if not playable:
            raise HTTPException(status_code=404, detail="Playable not found")
        
        # CRITICAL: Save skip progress SYNCHRONOUSLY to prevent duplicates in feed
        try:
            progress = {
                "user_id": current_user.user_id,
                "playable_id": playable_id,
                "answered": False,
                "skipped": True,
                "correct": False,
                "timestamp": datetime.now(timezone.utc)
            }
            await db.user_progress.insert_one(progress)
            logging.info(f"Skip saved (sync): user={current_user.user_id}, playable={playable_id}")
        except Exception as e:
            # Log but don't fail - duplicate key error is OK
            logging.warning(f"Skip save warning: {e}")
        
        # Get current user stats for response
        user_doc = await db.users.find_one({"user_id": current_user.user_id}, {"_id": 0})
        current_streak = user_doc.get("current_streak", 0)
        current_skipped = user_doc.get("skipped", 0)
        
        # ASYNC: Update skip count (non-critical)
        async def update_skip_count():
            try:
                await db.users.update_one(
                    {"user_id": current_user.user_id},
                    {"$inc": {"skipped": 1}}
                )
            except Exception as e:
                logging.error(f"Failed to update skip count: {e}")
        
        await task_manager.create_task(
            update_skip_count(),
            task_name=f"update_skip_count:{current_user.user_id}"
        )
        
        # Return immediately
        return {
            "skipped": True,
            "playable_id": playable_id,
            "current_streak": current_streak,
            "total_skipped": current_skipped + 1  # Optimistic update
        }
    
    except HTTPException:
        raise
    except Exception as e:
        logging.error(f"Error skipping playable: {e}")
        raise HTTPException(status_code=500, detail=str(e))

# ==================== SEED DATA ====================

@api_router.post("/seed")
async def seed_data():
    """Seed database with sample playables"""
    try:
        # Check if already seeded
        count = await db.playables.count_documents({})
        if count > 0:
            return {"message": "Database already seeded", "count": count}
        
        playables = [
            # Video + MCQ
            {
                "playable_id": f"play_{uuid.uuid4().hex[:12]}",
                "type": "video",
                "answer_type": "mcq",
                "category": "Science",
                "title": "How does photosynthesis work?",
                "question": {
                    "video_url": "https://www.w3schools.com/html/mov_bbb.mp4",
                    "text": "What is the primary gas absorbed during photosynthesis?"
                },
                "options": ["Oxygen", "Carbon Dioxide", "Nitrogen", "Hydrogen"],
                "correct_answer": "Carbon Dioxide",
                "difficulty": "easy",
                "created_at": datetime.now(timezone.utc)
            },
            # Text + MCQ
            {
                "playable_id": f"play_{uuid.uuid4().hex[:12]}",
                "type": "text",
                "answer_type": "mcq",
                "category": "History",
                "title": "World War II",
                "question": {
                    "text": "In which year did World War II end?"
                },
                "options": ["1943", "1944", "1945", "1946"],
                "correct_answer": "1945",
                "difficulty": "easy",
                "created_at": datetime.now(timezone.utc)
            },
            # Image + Text Input
            {
                "playable_id": f"play_{uuid.uuid4().hex[:12]}",
                "type": "image",
                "answer_type": "text_input",
                "category": "Geography",
                "title": "World Capitals",
                "question": {
                    "image_base64": "https://images.unsplash.com/photo-1511739001486-6bfe10ce785f?w=400&h=300&fit=crop",
                    "text": "Which city is this landmark located in?"
                },
                "options": None,
                "correct_answer": "Paris",
                "difficulty": "easy",
                "created_at": datetime.now(timezone.utc)
            },
            # Video + Text + MCQ
            {
                "playable_id": f"play_{uuid.uuid4().hex[:12]}",
                "type": "video_text",
                "answer_type": "mcq",
                "category": "Technology",
                "title": "Programming Basics",
                "question": {
                    "video_url": "https://www.w3schools.com/html/movie.mp4",
                    "text": "What does HTML stand for?"
                },
                "options": ["Hyper Text Markup Language", "High Tech Modern Language", "Home Tool Markup Language", "Hyperlinks and Text Markup Language"],
                "correct_answer": "Hyper Text Markup Language",
                "difficulty": "easy",
                "created_at": datetime.now(timezone.utc)
            },
            # Image + Text + MCQ
            {
                "playable_id": f"play_{uuid.uuid4().hex[:12]}",
                "type": "image_text",
                "answer_type": "mcq",
                "category": "Math",
                "title": "Basic Arithmetic",
                "question": {
                    "image_base64": "data:image/svg+xml;base64,PHN2ZyB3aWR0aD0iNDAwIiBoZWlnaHQ9IjMwMCIgeG1sbnM9Imh0dHA6Ly93d3cudzMub3JnLzIwMDAvc3ZnIj48cmVjdCB3aWR0aD0iNDAwIiBoZWlnaHQ9IjMwMCIgZmlsbD0iI0ZGRjdFRCIvPjx0ZXh0IHg9IjUwJSIgeT0iNTAlIiBmb250LXNpemU9IjYwIiBmaWxsPSIjMzMzIiB0ZXh0LWFuY2hvcj0ibWlkZGxlIiBkeT0iLjNlbSI+NSArIDcgPSA/PC90ZXh0Pjwvc3ZnPg==",
                    "text": "What is the answer to this equation?"
                },
                "options": ["10", "11", "12", "13"],
                "correct_answer": "12",
                "difficulty": "easy",
                "created_at": datetime.now(timezone.utc)
            },
            # Text + Text Input
            {
                "playable_id": f"play_{uuid.uuid4().hex[:12]}",
                "type": "text",
                "answer_type": "text_input",
                "category": "Literature",
                "title": "Famous Authors",
                "question": {
                    "text": "Who wrote 'Romeo and Juliet'?"
                },
                "options": None,
                "correct_answer": "Shakespeare",
                "difficulty": "easy",
                "created_at": datetime.now(timezone.utc)
            },
            # Video + Text Input
            {
                "playable_id": f"play_{uuid.uuid4().hex[:12]}",
                "type": "video",
                "answer_type": "text_input",
                "category": "Music",
                "title": "Musical Instruments",
                "question": {
                    "video_url": "https://www.w3schools.com/html/mov_bbb.mp4",
                    "text": "What instrument family does the piano belong to?"
                },
                "options": None,
                "correct_answer": "Percussion",
                "difficulty": "medium",
                "created_at": datetime.now(timezone.utc)
            },
            # Image + Text + Text Input
            {
                "playable_id": f"play_{uuid.uuid4().hex[:12]}",
                "type": "image_text",
                "answer_type": "text_input",
                "category": "Art",
                "title": "Famous Paintings",
                "question": {
                    "image_base64": "https://images.unsplash.com/photo-1574870111867-089730e5a72b?w=400&h=300&fit=crop",
                    "text": "Who painted this famous artwork?"
                },
                "options": None,
                "correct_answer": "Leonardo da Vinci",
                "difficulty": "medium",
                "created_at": datetime.now(timezone.utc)
            }
        ]
        
        await db.playables.insert_many(playables)
        
        return {"message": "Database seeded successfully", "count": len(playables)}
    
    except Exception as e:
        logging.error(f"Error seeding data: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@api_router.delete("/user/reset-progress")
async def reset_user_progress(current_user: User = Depends(require_auth)):
    """Reset current user's progress (for Play Again feature)
    
    This will:
    1. Delete all entries from user_progress collection
    2. Reset all stats in users collection to zero
    """
    try:
        # Delete ALL user's progress entries
        deleted = await db.user_progress.delete_many({"user_id": current_user.user_id})
        
        # Reset ALL user stats to zero
        await db.users.update_one(
            {"user_id": current_user.user_id},
            {"$set": {
                "total_played": 0,
                "correct_answers": 0,
                "current_streak": 0,
                "best_streak": 0,
                "skipped": 0
            }}
        )
        
        return {
            "message": "Progress reset successfully",
            "deleted_progress_count": deleted.deleted_count
        }
    
    except Exception as e:
        logging.error(f"Error resetting progress: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@api_router.post("/dev/reset-progress")
async def dev_reset_progress(email: str = "supanshah51191@gmail.com"):
    """Reset user progress by email (dev only)"""
    try:
        # Find user by email
        user = await db.users.find_one({"email": email})
        if not user:
            return {"message": f"User {email} not found"}
        
        user_id = user.get("user_id", str(user.get("_id")))
        
        # Delete user's progress
        await db.user_progress.delete_many({"user_id": user_id})
        
        # Reset user stats
        await db.users.update_one(
            {"email": email},
            {"$set": {
                "total_played": 0,
                "correct_answers": 0,
                "current_streak": 0,
                "best_streak": 0
            }}
        )
        
        return {"message": f"Progress reset for {email}"}
    
    except Exception as e:
        logging.error(f"Error resetting progress: {e}")
        raise HTTPException(status_code=500, detail=str(e))

# ==================== ADMIN ENDPOINTS ====================

# Admin credentials (in production, use environment variables)
# Admin credentials - multiple admins supported
ADMIN_USERS = {
    "admin": "@dm!n!spl@ying",
    "meenal": "M3en@ladmin",
    "parul": "P@rul0ps"
}

class AdminLoginRequest(BaseModel):
    username: str
    password: str

class AdminResetProgressRequest(BaseModel):
    email: str

class AddPlayableRequest(BaseModel):
    type: str  # "text", "image_text", "video_text", "guess_the_x", "chess_mate_in_2", "this_or_that", "wordle"
    answer_type: Optional[str] = None  # "mcq", "text_input", "tap_select", "wordle_grid" - auto-set for special types
    category: str
    title: Optional[str] = None  # Deprecated - kept for backward compatibility
    question_text: Optional[str] = None
    image_url: Optional[str] = None
    video_url: Optional[str] = None
    video_start: Optional[int] = None  # Start time in seconds
    video_end: Optional[int] = None    # End time in seconds
    options: Optional[List[str]] = None  # For MCQ (4 options)
    correct_answer: str
    alternate_answers: Optional[List[str]] = None  # For text_input: spelling variants, short forms
    answer_explanation: Optional[str] = None  # Brief explanation of the answer
    hints: Optional[List[str]] = None  # For guess_the_x: 3-5 hints
    fen: Optional[str] = None  # For chess_mate_in_2: FEN position string
    solution: Optional[List[str]] = None  # For chess_mate_in_2: ALL moves in UCI format
    # This or That fields
    image_left_url: Optional[str] = None  # Left image URL
    image_right_url: Optional[str] = None  # Right image URL
    label_left: Optional[str] = None  # Label for left image (used for answer matching)
    label_right: Optional[str] = None  # Label for right image (used for answer matching)
    difficulty: str = "medium"
    weight: int = 0  # Ranking weight: 0 or positive integer. Higher = shown first
    status: str = "active"  # Status: "active" or "inactive"

@api_router.post("/admin/login")
async def admin_login(request: AdminLoginRequest):
    """Admin login endpoint"""
    # Check if username exists and password matches
    if request.username in ADMIN_USERS and ADMIN_USERS[request.username] == request.password:
        # Generate admin session token
        admin_token = f"admin_{uuid.uuid4().hex}"
        
        try:
            # Store admin session (expires in 24 hours)
            result = await db.admin_sessions.insert_one({
                "token": admin_token,
                "username": request.username,  # Track which admin logged in
                "created_at": datetime.now(timezone.utc),
                "expires_at": datetime.now(timezone.utc) + timedelta(hours=24)
            })
            logging.info(f"Admin session created for {request.username}: {admin_token}, inserted_id: {result.inserted_id}")
        except Exception as e:
            logging.error(f"Failed to create admin session: {e}")
            raise HTTPException(status_code=500, detail="Failed to create session")
        
        return {"success": True, "token": admin_token}
    else:
        raise HTTPException(status_code=401, detail="Invalid credentials")

async def verify_admin_token(authorization: Optional[str] = Header(None)):
    """Verify admin token"""
    logging.info(f"verify_admin_token called with auth: {authorization}")
    if not authorization:
        logging.warning("No authorization header provided")
        raise HTTPException(status_code=401, detail="Admin token required")
    
    token = authorization.replace("Bearer ", "")
    logging.info(f"Verifying admin token: {token[:20]}...")
    
    try:
        session = await db.admin_sessions.find_one({"token": token})
        logging.info(f"Session lookup result: {session}")
    except Exception as e:
        logging.error(f"Error looking up session: {e}")
        raise HTTPException(status_code=500, detail="Database error")
    
    if not session:
        logging.warning(f"Admin session not found for token: {token[:20]}...")
        raise HTTPException(status_code=401, detail="Invalid admin token - session not found")
    
    logging.info(f"Admin session found, expires at: {session.get('expires_at')}")
    
    # Check expiry
    expires_at = session["expires_at"]
    if expires_at.tzinfo is None:
        expires_at = expires_at.replace(tzinfo=timezone.utc)
    
    if expires_at < datetime.now(timezone.utc):
        raise HTTPException(status_code=401, detail="Admin session expired")
    
    return True

@api_router.post("/admin/reset-user-progress")
async def admin_reset_user_progress(
    request: AdminResetProgressRequest,
    _: bool = Depends(verify_admin_token)
):
    """Reset progress for a specific user by email (admin only)
    
    This will:
    1. Delete all entries from user_progress collection
    2. Reset all stats in users collection to zero
    """
    try:
        # Find user by email
        user = await db.users.find_one({"email": request.email})
        if not user:
            raise HTTPException(status_code=404, detail=f"User with email {request.email} not found")
        
        user_id = user.get("user_id", str(user.get("_id")))
        
        # Delete ALL user's progress entries (both answered and skipped)
        deleted = await db.user_progress.delete_many({"user_id": user_id})
        
        # Reset ALL user stats to zero
        await db.users.update_one(
            {"email": request.email},
            {"$set": {
                "total_played": 0,
                "correct_answers": 0,
                "current_streak": 0,
                "best_streak": 0,
                "skipped": 0
            }}
        )
        
        return {
            "success": True,
            "message": f"Progress fully reset for {request.email}",
            "deleted_progress_count": deleted.deleted_count,
            "stats_reset": ["total_played", "correct_answers", "current_streak", "best_streak", "skipped"]
        }
    
    except HTTPException:
        raise
    except Exception as e:
        logging.error(f"Error resetting progress: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@api_router.post("/admin/add-playable")
async def admin_add_playable(
    request: AddPlayableRequest,
    _: bool = Depends(verify_admin_token)
):
    """Add a new playable content (admin only)"""
    try:
        # Validate category exists
        category_exists = await db.categories.find_one({"name": {"$regex": f"^{request.category}$", "$options": "i"}})
        if not category_exists:
            raise HTTPException(status_code=400, detail=f"Category '{request.category}' does not exist. Please add it first in the Categories tab.")
        
        # Use the exact category name from database (preserves case)
        category_name = category_exists["name"]
        
        # Validate based on type
        question = {}
        
        # Add text if provided
        if request.question_text:
            question["text"] = request.question_text
        
        # Add image URL if provided
        if request.image_url:
            question["image_base64"] = request.image_url  # Using same field name for compatibility
        
        # Add video URL if provided
        if request.video_url:
            question["video_url"] = request.video_url
        
        # Validate type is one of the supported formats
        if request.type not in PLAYABLE_TYPE_CONFIG:
            supported = list(PLAYABLE_TYPE_CONFIG.keys())
            raise HTTPException(
                status_code=400, 
                detail=f"Unsupported type '{request.type}'. Supported types: {', '.join(supported)}"
            )
        
        # Get valid answer_type based on playable type (auto-determine or validate)
        try:
            answer_type = get_valid_answer_type(request.type, request.answer_type)
        except ValueError as e:
            raise HTTPException(status_code=400, detail=str(e))
        
        # Validate answer_type if user provided one that doesn't match
        type_config = PLAYABLE_TYPE_CONFIG[request.type]
        if request.answer_type and request.answer_type not in type_config["valid_answer_types"]:
            valid = type_config["valid_answer_types"]
            raise HTTPException(
                status_code=400,
                detail=f"Invalid answer_type '{request.answer_type}' for type '{request.type}'. Valid options: {', '.join(valid)}"
            )
        
        # Validate required fields based on type
        if request.type == "text" and not request.question_text:
            raise HTTPException(status_code=400, detail="Text question requires question_text")
        
        if request.type == "image_text" and not request.image_url:
            raise HTTPException(status_code=400, detail="Image question requires image_url")
        
        if request.type == "video_text" and not request.video_url:
            raise HTTPException(status_code=400, detail="Video question requires video_url")
        
        # Validate guess_the_x has hints
        if request.type == "guess_the_x":
            if not request.hints or len(request.hints) < 3:
                raise HTTPException(status_code=400, detail="Guess the X requires at least 3 hints")
            if len(request.hints) > 5:
                raise HTTPException(status_code=400, detail="Guess the X allows maximum 5 hints")
        
        # Validate chess_mate_in_2 has FEN and solution
        if request.type == "chess_mate_in_2":
            if not request.fen:
                raise HTTPException(status_code=400, detail="Chess puzzle requires a FEN position")
            if not request.solution or len(request.solution) < 1:
                raise HTTPException(status_code=400, detail="Chess puzzle requires at least 1 solution move")
        
        # Validate this_or_that has both images and labels
        if request.type == "this_or_that":
            if not request.image_left_url or not request.image_right_url:
                raise HTTPException(status_code=400, detail="This or That requires both left and right images")
            if not request.label_left or not request.label_right:
                raise HTTPException(status_code=400, detail="This or That requires labels for both images")
            if request.correct_answer not in [request.label_left, request.label_right]:
                raise HTTPException(status_code=400, detail="Correct answer must match one of the labels")
        
        # Validate wordle has 5-letter word
        if request.type == "wordle":
            if not request.correct_answer or len(request.correct_answer) != 5:
                raise HTTPException(status_code=400, detail="Wordle requires a 5-letter word as correct_answer")
            if not request.correct_answer.isalpha():
                raise HTTPException(status_code=400, detail="Wordle word must contain only letters")
        
        if request.answer_type == "mcq" and request.type not in ["guess_the_x", "chess_mate_in_2", "this_or_that"]:
            if not request.options or len(request.options) < 2:
                raise HTTPException(status_code=400, detail="MCQ requires at least 2 options")
            if request.correct_answer not in request.options:
                raise HTTPException(status_code=400, detail="Correct answer must be one of the options")
        
        # Build question object for this_or_that
        if request.type == "this_or_that":
            question = {
                "text": request.question_text,
                "image_left": request.image_left_url,
                "image_right": request.image_right_url,
                "label_left": request.label_left,
                "label_right": request.label_right
            }
        
        # Create playable
        playable_id = f"play_{uuid.uuid4().hex[:12]}"
        playable_doc = {
            "playable_id": playable_id,
            "type": request.type,
            "answer_type": "tap_select" if request.type == "this_or_that" else ("text_input" if request.type in ["guess_the_x", "chess_mate_in_2"] else request.answer_type),
            "category": request.category,
            "question": question,
            "options": request.options if request.answer_type == "mcq" and request.type not in ["guess_the_x", "chess_mate_in_2", "this_or_that", "wordle"] else None,
            "correct_answer": request.correct_answer,
            "alternate_answers": request.alternate_answers if (request.answer_type == "text_input" or request.type in ["guess_the_x", "chess_mate_in_2"]) else None,
            "answer_explanation": request.answer_explanation,
            "hints": request.hints if request.type == "guess_the_x" else None,
            "fen": request.fen if request.type == "chess_mate_in_2" else None,
            "solution": request.solution if request.type == "chess_mate_in_2" else None,
            "video_start": request.video_start if request.type in ["video", "video_text"] else None,
            "video_end": request.video_end if request.type in ["video", "video_text"] else None,
            "difficulty": request.difficulty,
            "weight": max(0, request.weight),  # Ensure weight is 0 or positive
            "status": request.status,  # active or inactive
            "created_at": datetime.now(timezone.utc)
        }
        
        # Use the validated answer_type (auto-determined for special types)
        playable_doc["answer_type"] = answer_type
        
        # Use the validated category name
        playable_doc["category"] = category_name
        
        await db.playables.insert_one(playable_doc)
        
        return {
            "success": True,
            "message": "Playable added successfully",
            "playable_id": playable_id
        }
    
    except HTTPException:
        raise
    except Exception as e:
        logging.error(f"Error adding playable: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@api_router.post("/admin/migrate/add-status-field")
async def migrate_add_status_field(_: bool = Depends(verify_admin_token)):
    """
    One-time migration: Add status='active' to all playables that don't have a status field.
    Safe to run multiple times - only updates documents without status field.
    """
    try:
        # Count documents without status
        without_status = await db.playables.count_documents({"status": {"$exists": False}})
        
        if without_status == 0:
            return {
                "success": True,
                "message": "No migration needed - all playables already have status field",
                "updated": 0
            }
        
        # Update all playables without status to have status="active"
        result = await db.playables.update_many(
            {"status": {"$exists": False}},
            {"$set": {"status": "active"}}
        )
        
        # Create index on status field
        try:
            await db.playables.create_index("status")
        except Exception:
            pass  # Index may already exist
        
        # Get counts for verification
        total = await db.playables.count_documents({})
        active = await db.playables.count_documents({"status": "active"})
        inactive = await db.playables.count_documents({"status": "inactive"})
        
        return {
            "success": True,
            "message": f"Migration complete! Updated {result.modified_count} playables",
            "updated": result.modified_count,
            "total": total,
            "active": active,
            "inactive": inactive
        }
    except Exception as e:
        logging.error(f"Migration error: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@api_router.post("/admin/migrate/set-status-by-answer-type")
async def migrate_set_status_by_answer_type(
    answer_type: str = Query(..., description="Answer type to filter (e.g., text_input, mcq)"),
    new_status: str = Query(..., description="New status to set (active or inactive)"),
    _: bool = Depends(verify_admin_token)
):
    """
    Set status for all playables matching a specific answer_type.
    Example: Set all text_input (typing) questions to inactive.
    """
    try:
        # Validate new_status
        if new_status not in ["active", "inactive"]:
            raise HTTPException(status_code=400, detail="status must be 'active' or 'inactive'")
        
        # Count matching documents
        match_count = await db.playables.count_documents({"answer_type": answer_type})
        
        if match_count == 0:
            return {
                "success": True,
                "message": f"No playables found with answer_type='{answer_type}'",
                "updated": 0
            }
        
        # Update all matching playables
        result = await db.playables.update_many(
            {"answer_type": answer_type},
            {"$set": {"status": new_status}}
        )
        
        # Get counts for verification
        total = await db.playables.count_documents({})
        active = await db.playables.count_documents({"status": "active"})
        inactive = await db.playables.count_documents({"status": "inactive"})
        
        return {
            "success": True,
            "message": f"Updated {result.modified_count} playables with answer_type='{answer_type}' to status='{new_status}'",
            "updated": result.modified_count,
            "matched": match_count,
            "total": total,
            "active": active,
            "inactive": inactive
        }
    except HTTPException:
        raise
    except Exception as e:
        logging.error(f"Migration error: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@api_router.post("/admin/migrate/fix-answer-types")
async def migrate_fix_answer_types(_: bool = Depends(verify_admin_token)):
    """
    Fix answer_type for playables that have incorrect values.
    Updates answer_type based on PLAYABLE_TYPE_CONFIG.
    """
    try:
        results = {}
        
        # Fix each type that has auto-determined answer_type
        for playable_type, config in PLAYABLE_TYPE_CONFIG.items():
            if len(config["valid_answer_types"]) == 1:
                correct_answer_type = config["valid_answer_types"][0]
                
                # Find playables with wrong answer_type
                wrong_count = await db.playables.count_documents({
                    "type": playable_type,
                    "answer_type": {"$ne": correct_answer_type}
                })
                
                if wrong_count > 0:
                    result = await db.playables.update_many(
                        {"type": playable_type, "answer_type": {"$ne": correct_answer_type}},
                        {"$set": {"answer_type": correct_answer_type}}
                    )
                    results[playable_type] = {
                        "fixed": result.modified_count,
                        "correct_answer_type": correct_answer_type
                    }
        
        return {
            "success": True,
            "message": "Answer types fixed",
            "results": results
        }
    except Exception as e:
        logging.error(f"Migration error: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@api_router.get("/admin/playables")
async def admin_get_playables(
    _: bool = Depends(verify_admin_token),
    page: int = Query(1, ge=1, description="Page number"),
    limit: int = Query(100, ge=1, le=500, description="Items per page"),
    category: Optional[str] = Query(None, description="Filter by category"),
    type: Optional[str] = Query(None, description="Filter by playable type"),
    status: Optional[str] = Query("active", description="Filter by status: active, inactive, or all")
):
    """Get all playables with pagination (admin only)"""
    try:
        # Build filter
        filter_query = {}
        if category:
            filter_query["category"] = category
        if type:
            filter_query["type"] = type
        
        # Status filter - default shows active only
        if status == "active":
            filter_query["$or"] = [
                {"status": "active"},
                {"status": {"$exists": False}}  # Legacy playables without status field
            ]
        elif status == "inactive":
            filter_query["status"] = "inactive"
        # If status == "all", don't add any status filter
        
        # Get total count
        total_count = await db.playables.count_documents(filter_query)
        
        # Calculate skip
        skip = (page - 1) * limit
        
        # Get paginated playables
        playables = await db.playables.find(filter_query, {"_id": 0}).sort("created_at", -1).skip(skip).limit(limit).to_list(limit)
        
        return {
            "playables": playables,
            "count": len(playables),
            "total": total_count,
            "page": page,
            "limit": limit,
            "total_pages": (total_count + limit - 1) // limit,
            "status_filter": status
        }
    except Exception as e:
        logging.error(f"Error getting playables: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@api_router.delete("/admin/playables/{playable_id}")
async def admin_delete_playable(playable_id: str, _: bool = Depends(verify_admin_token)):
    """Delete a playable (admin only)"""
    try:
        result = await db.playables.delete_one({"playable_id": playable_id})
        if result.deleted_count == 0:
            raise HTTPException(status_code=404, detail="Playable not found")
        return {"success": True, "message": "Playable deleted"}
    except HTTPException:
        raise
    except Exception as e:
        logging.error(f"Error deleting playable: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@api_router.put("/admin/playables/{playable_id}")
async def admin_update_playable(playable_id: str, request: AddPlayableRequest, _: bool = Depends(verify_admin_token)):
    """Update a playable (admin only)"""
    try:
        # Check if playable exists
        existing = await db.playables.find_one({"playable_id": playable_id})
        if not existing:
            raise HTTPException(status_code=404, detail="Playable not found")
        
        # Build question object - MERGE with existing question data
        existing_question = existing.get("question", {})
        question = existing_question.copy()  # Start with existing data
        
        # Update only the fields that are provided
        if request.question_text is not None:
            question["text"] = request.question_text
        if request.image_url is not None:
            if request.image_url.startswith("data:"):
                question["image_base64"] = request.image_url
                question.pop("image_url", None)  # Remove other image field
            else:
                question["image_url"] = request.image_url
                question.pop("image_base64", None)  # Remove other image field
        if request.video_url is not None:
            question["video_url"] = request.video_url
        
        # Build update document
        update_doc = {
            "type": request.type,
            "answer_type": request.answer_type,
            "category": request.category,
            "question": question,
            "options": request.options if request.answer_type == "mcq" and request.type not in ["guess_the_x", "chess_mate_in_2"] else None,
            "correct_answer": request.correct_answer,
            "alternate_answers": request.alternate_answers if (request.answer_type == "text_input" or request.type in ["guess_the_x", "chess_mate_in_2"]) else None,
            "answer_explanation": request.answer_explanation,
            "hints": request.hints if request.type == "guess_the_x" else None,
            "fen": request.fen if request.type == "chess_mate_in_2" else None,
            "solution": request.solution if request.type == "chess_mate_in_2" else None,
            "video_start": request.video_start,
            "video_end": request.video_end,
            "difficulty": request.difficulty,
            "weight": max(0, request.weight),  # Ensure weight is 0 or positive
            "updated_at": datetime.now(timezone.utc)
        }
        
        await db.playables.update_one(
            {"playable_id": playable_id},
            {"$set": update_doc}
        )
        
        return {"success": True, "message": "Playable updated successfully", "playable_id": playable_id}
    except HTTPException:
        raise
    except Exception as e:
        logging.error(f"Error updating playable: {e}")
        raise HTTPException(status_code=500, detail=str(e))

class PartialUpdateRequest(BaseModel):
    """Partial update request - for updating weight and difficulty only"""
    weight: Optional[int] = None
    difficulty: Optional[str] = None

@api_router.patch("/admin/playables/{playable_id}")
async def admin_patch_playable(playable_id: str, request: PartialUpdateRequest, _: bool = Depends(verify_admin_token)):
    """Partially update a playable - only updates provided fields (admin only)"""
    try:
        # Check if playable exists
        existing = await db.playables.find_one({"playable_id": playable_id})
        if not existing:
            raise HTTPException(status_code=404, detail="Playable not found")
        
        # Build update document with only provided fields
        update_doc = {}
        
        if request.weight is not None:
            update_doc["weight"] = max(0, request.weight)
        if request.difficulty is not None:
            update_doc["difficulty"] = request.difficulty
        
        if not update_doc:
            raise HTTPException(status_code=400, detail="No fields provided for update")
        
        update_doc["updated_at"] = datetime.now(timezone.utc)
        
        await db.playables.update_one(
            {"playable_id": playable_id},
            {"$set": update_doc}
        )
        
        return {"success": True, "message": "Playable updated successfully", "playable_id": playable_id}
    except HTTPException:
        raise
    except Exception as e:
        logging.error(f"Error patching playable: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@api_router.post("/admin/remove-titles")
async def admin_remove_titles(_: bool = Depends(verify_admin_token)):
    """Remove title field from all playables (admin only) - one-time migration"""
    try:
        # Count before
        count_before = await db.playables.count_documents({"title": {"$exists": True}})
        
        # Remove title field from all playables
        result = await db.playables.update_many(
            {},
            {"$unset": {"title": ""}}
        )
        
        # Verify
        count_after = await db.playables.count_documents({"title": {"$exists": True}})
        
        return {
            "success": True,
            "message": "Title field removed from all playables",
            "playables_with_title_before": count_before,
            "playables_modified": result.modified_count,
            "playables_with_title_after": count_after
        }
    except Exception as e:
        logging.error(f"Error removing titles: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@api_router.get("/admin/users")
async def admin_get_users(_: bool = Depends(verify_admin_token)):
    """Get all users (admin only)"""
    try:
        users = await db.users.find({}, {"_id": 0}).sort("created_at", -1).to_list(100)
        return {"users": users, "count": len(users)}
    except Exception as e:
        logging.error(f"Error getting users: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@api_router.get("/admin/user-progress/{email}")
async def admin_get_user_progress(email: str, _: bool = Depends(verify_admin_token)):
    """Get user progress records for debugging (admin only)"""
    try:
        # Find user by email
        user = await db.users.find_one({"email": email})
        if not user:
            raise HTTPException(status_code=404, detail=f"User with email {email} not found")
        
        user_id = user.get("user_id")
        
        # Get all progress records
        progress_records = await db.user_progress.find(
            {"user_id": user_id}
        ).to_list(1000)
        
        # Get all playables for reference
        all_playables = await db.playables.find({}, {"playable_id": 1, "title": 1, "weight": 1}).to_list(100)
        playable_map = {p["playable_id"]: p for p in all_playables}
        
        # Enrich progress with playable info
        enriched_progress = []
        for p in progress_records:
            playable_info = playable_map.get(p["playable_id"], {})
            enriched_progress.append({
                "playable_id": p["playable_id"],
                "title": playable_info.get("title", "UNKNOWN"),
                "weight": playable_info.get("weight", 0),
                "answered": p.get("answered", False),
                "skipped": p.get("skipped", False),
                "correct": p.get("correct", False),
                "timestamp": str(p.get("timestamp", ""))
            })
        
        # Sort by timestamp
        enriched_progress.sort(key=lambda x: x["timestamp"])
        
        # Find which playables are NOT in progress (remaining)
        progress_ids = {p["playable_id"] for p in progress_records}
        remaining = [p for p in all_playables if p["playable_id"] not in progress_ids]
        remaining.sort(key=lambda x: x.get("weight", 0), reverse=True)
        
        return {
            "email": email,
            "user_id": user_id,
            "total_playables": len(all_playables),
            "progress_count": len(progress_records),
            "remaining_count": len(remaining),
            "progress": enriched_progress,
            "remaining_playables": [
                {"playable_id": p["playable_id"], "title": p.get("title"), "weight": p.get("weight", 0)}
                for p in remaining
            ]
        }
    except HTTPException:
        raise
    except Exception as e:
        logging.error(f"Error getting user progress: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@api_router.get("/admin/task-status")
async def admin_get_task_status(_: bool = Depends(verify_admin_token)):
    """Get background task manager status (admin only)
    
    Returns:
    - pending_count: Number of tasks currently queued
    - is_shutdown: Whether shutdown has been initiated
    """
    return {
        "pending_count": task_manager.pending_count,
        "is_shutdown": task_manager._shutdown,
        "max_retries": task_manager._max_retries,
        "retry_delay": task_manager._retry_delay
    }

@api_router.post("/admin/reconcile-user-stats")
async def admin_reconcile_user_stats(request: dict, _: bool = Depends(verify_admin_token)):
    """Reconcile user stats from user_progress records (admin only)
    
    This fixes any mismatch between user_progress records and users.total_played
    """
    try:
        email = request.get("email")
        if not email:
            raise HTTPException(status_code=400, detail="Email is required")
        
        # Find user
        user = await db.users.find_one({"email": email})
        if not user:
            raise HTTPException(status_code=404, detail=f"User with email {email} not found")
        
        user_id = user["user_id"]
        
        # Count actual progress records (excluding skipped)
        answered_records = await db.user_progress.find({
            "user_id": user_id,
            "answered": True
        }).to_list(10000)
        
        total_played = len(answered_records)
        correct_answers = sum(1 for r in answered_records if r.get("correct", False))
        
        # Get skipped count
        skipped_records = await db.user_progress.find({
            "user_id": user_id,
            "skipped": True
        }).to_list(10000)
        skipped_count = len(skipped_records)
        
        # Update user record
        old_stats = {
            "total_played": user.get("total_played", 0),
            "correct_answers": user.get("correct_answers", 0),
            "skipped": user.get("skipped", 0)
        }
        
        await db.users.update_one(
            {"user_id": user_id},
            {"$set": {
                "total_played": total_played,
                "correct_answers": correct_answers,
                "skipped": skipped_count
            }}
        )
        
        return {
            "success": True,
            "email": email,
            "user_id": user_id,
            "old_stats": old_stats,
            "new_stats": {
                "total_played": total_played,
                "correct_answers": correct_answers,
                "skipped": skipped_count
            },
            "message": f"Reconciled stats: {total_played} played, {correct_answers} correct, {skipped_count} skipped"
        }
    except HTTPException:
        raise
    except Exception as e:
        logging.error(f"Error reconciling user stats: {e}")
        raise HTTPException(status_code=500, detail=str(e))

# ==================== ADMIN CATEGORY MANAGEMENT ====================

@api_router.get("/admin/categories")
async def admin_get_categories(_: bool = Depends(verify_admin_token)):
    """Get all managed categories (admin only)"""
    try:
        categories = await db.categories.find({}, {"_id": 0}).sort("name", 1).to_list(100)
        
        # Get playable counts for each category
        pipeline = [
            {"$group": {"_id": "$category", "count": {"$sum": 1}}}
        ]
        category_counts = await db.playables.aggregate(pipeline).to_list(100)
        count_map = {c["_id"]: c["count"] for c in category_counts}
        
        for cat in categories:
            cat["playable_count"] = count_map.get(cat["name"], 0)
        
        return {"categories": categories, "count": len(categories)}
    except Exception as e:
        logging.error(f"Error getting categories: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@api_router.get("/admin/valid-icons")
async def get_valid_icons(_: bool = Depends(verify_admin_token)):
    """Get list of all valid Ionicons names for category icons"""
    return {
        "icons": sorted(list(VALID_IONICONS)),
        "count": len(VALID_IONICONS)
    }

@api_router.post("/admin/categories")
async def admin_add_category(request: AddCategoryRequest, _: bool = Depends(verify_admin_token)):
    """Add a new category (admin only)"""
    try:
        # Validate icon name
        icon_to_use = request.icon or "help-circle"
        if not is_valid_icon(icon_to_use):
            raise HTTPException(
                status_code=400, 
                detail=f"Invalid icon '{icon_to_use}'. Please use a valid Ionicons name (e.g., 'flask', 'globe', 'star', 'rocket')"
            )
        
        # Check if category already exists (case-insensitive)
        existing = await db.categories.find_one({"name": {"$regex": f"^{request.name}$", "$options": "i"}})
        if existing:
            raise HTTPException(status_code=400, detail=f"Category '{request.name}' already exists")
        
        category_id = f"cat_{uuid.uuid4().hex[:8]}"
        category_doc = {
            "category_id": category_id,
            "name": request.name,
            "icon": icon_to_use,
            "color": request.color or "#00FF87",
            "created_at": datetime.now(timezone.utc)
        }
        
        # Add description if provided
        if request.description:
            category_doc["description"] = request.description
        
        await db.categories.insert_one(category_doc)
        
        return {
            "success": True,
            "message": f"Category '{request.name}' added successfully",
            "category_id": category_id
        }
    except HTTPException:
        raise
    except Exception as e:
        logging.error(f"Error adding category: {e}")
        raise HTTPException(status_code=500, detail=str(e))

# Model for updating a category
class UpdateCategoryRequest(BaseModel):
    icon: Optional[str] = None
    color: Optional[str] = None
    description: Optional[str] = None  # Optional description for the category

@api_router.patch("/admin/categories/{category_id}")
async def admin_update_category(
    category_id: str, 
    request: UpdateCategoryRequest, 
    _: bool = Depends(verify_admin_token)
):
    """Update a category's icon and/or color (admin only)"""
    try:
        # Find category
        category = await db.categories.find_one({"category_id": category_id})
        if not category:
            raise HTTPException(status_code=404, detail="Category not found")
        
        # Build update dict
        update_fields = {}
        
        if request.icon is not None:
            # Validate icon
            if not is_valid_icon(request.icon):
                raise HTTPException(
                    status_code=400,
                    detail=f"Invalid icon '{request.icon}'. Please use a valid Ionicons name."
                )
            update_fields["icon"] = request.icon
        
        if request.color is not None:
            # Basic color validation (hex format)
            color = request.color.strip()
            if not color.startswith("#") or len(color) not in [4, 7]:
                raise HTTPException(
                    status_code=400,
                    detail=f"Invalid color '{color}'. Please use hex format (e.g., '#FF5722')"
                )
            update_fields["color"] = color
        
        # Handle description update (can be set to empty string to clear)
        if request.description is not None:
            if request.description.strip():
                update_fields["description"] = request.description.strip()
            else:
                # Empty string means remove the description
                update_fields["description"] = None
        
        if not update_fields:
            raise HTTPException(status_code=400, detail="No fields to update. Provide 'icon', 'color', and/or 'description'.")
        
        # Update category
        await db.categories.update_one(
            {"category_id": category_id},
            {"$set": update_fields}
        )
        
        return {
            "success": True,
            "message": f"Category '{category['name']}' updated successfully",
            "updated_fields": list(update_fields.keys())
        }
    except HTTPException:
        raise
    except Exception as e:
        logging.error(f"Error updating category: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@api_router.delete("/admin/categories/{category_id}")
async def admin_delete_category(category_id: str, _: bool = Depends(verify_admin_token)):
    """Delete a category (admin only) - only if no playables use it"""
    try:
        # Find category
        category = await db.categories.find_one({"category_id": category_id})
        if not category:
            raise HTTPException(status_code=404, detail="Category not found")
        
        # Check if any playables use this category
        playable_count = await db.playables.count_documents({"category": category["name"]})
        if playable_count > 0:
            raise HTTPException(
                status_code=400, 
                detail=f"Cannot delete category '{category['name']}' - it has {playable_count} playable(s) using it"
            )
        
        await db.categories.delete_one({"category_id": category_id})
        
        return {"success": True, "message": f"Category '{category['name']}' deleted"}
    except HTTPException:
        raise
    except Exception as e:
        logging.error(f"Error deleting category: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@api_router.post("/admin/categories/init")
async def admin_init_categories(_: bool = Depends(verify_admin_token)):
    """Initialize categories from existing playables (admin only)"""
    try:
        # Get distinct categories from playables
        distinct_categories = await db.playables.distinct("category")
        
        added = []
        skipped = []
        
        for cat_name in distinct_categories:
            if not cat_name:
                continue
            
            # Check if already exists
            existing = await db.categories.find_one({"name": {"$regex": f"^{cat_name}$", "$options": "i"}})
            if existing:
                skipped.append(cat_name)
                continue
            
            # Get default style
            style = get_default_category_style(cat_name)
            
            category_doc = {
                "category_id": f"cat_{uuid.uuid4().hex[:8]}",
                "name": cat_name,
                "icon": style["icon"],
                "color": style["color"],
                "created_at": datetime.now(timezone.utc)
            }
            
            await db.categories.insert_one(category_doc)
            added.append(cat_name)
        
        return {
            "success": True,
            "message": f"Initialized {len(added)} categories",
            "added": added,
            "skipped": skipped
        }
    except Exception as e:
        logging.error(f"Error initializing categories: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@api_router.post("/admin/categories/fix-icons")
async def admin_fix_category_icons(_: bool = Depends(verify_admin_token)):
    """Update all category icons to use the correct defaults (admin only)"""
    try:
        categories = await db.categories.find({}).to_list(100)
        updated = []
        
        for cat in categories:
            cat_name = cat.get("name", "")
            style = get_default_category_style(cat_name)
            
            # Update if icon is different or is help-circle (default)
            current_icon = cat.get("icon", "help-circle")
            new_icon = style["icon"]
            new_color = style["color"]
            
            if current_icon != new_icon or current_icon == "help-circle":
                await db.categories.update_one(
                    {"category_id": cat["category_id"]},
                    {"$set": {"icon": new_icon, "color": new_color}}
                )
                updated.append(f"{cat_name}: {current_icon} → {new_icon}")
        
        return {
            "success": True,
            "message": f"Updated {len(updated)} categories",
            "updated": updated
        }
    except Exception as e:
        logging.error(f"Error fixing category icons: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@api_router.get("/admin/stats")
async def admin_get_stats(date: str = None, _: bool = Depends(verify_admin_token)):
    """Get user performance stats for a specific date (admin only)
    
    Args:
        date: Date in YYYY-MM-DD format. If not provided, returns stats for today.
    """
    try:
        # Parse date or use today
        if date:
            try:
                target_date = datetime.strptime(date, "%Y-%m-%d").replace(tzinfo=timezone.utc)
            except ValueError:
                raise HTTPException(status_code=400, detail="Invalid date format. Use YYYY-MM-DD")
        else:
            target_date = datetime.now(timezone.utc).replace(hour=0, minute=0, second=0, microsecond=0)
        
        # Date range for the selected day
        start_of_day = target_date.replace(hour=0, minute=0, second=0, microsecond=0)
        end_of_day = start_of_day + timedelta(days=1)
        
        # Get all users
        all_users = await db.users.find({}, {"_id": 0}).to_list(1000)
        
        # Get progress records for the target date
        progress_records = await db.user_progress.find({
            "timestamp": {"$gte": start_of_day, "$lt": end_of_day}
        }, {"_id": 0}).to_list(10000)
        
        # Build stats per user
        user_stats = []
        for user in all_users:
            user_id = user.get("user_id")
            email = user.get("email", "Unknown")
            name = user.get("name", "Unknown")
            
            # Filter progress for this user on target date
            user_progress = [p for p in progress_records if p.get("user_id") == user_id]
            
            # Separate answered vs skipped
            answered_progress = [p for p in user_progress if p.get("answered", False)]
            skipped_progress = [p for p in user_progress if p.get("skipped", False)]
            
            played_count = len(answered_progress)
            skipped_count = len(skipped_progress)
            correct_count = sum(1 for p in answered_progress if p.get("correct", False))
            incorrect_count = played_count - correct_count
            
            # Calculate accuracy
            accuracy = round((correct_count / played_count * 100), 1) if played_count > 0 else 0
            
            user_stats.append({
                "user_id": user_id,
                "email": email,
                "name": name,
                "played": played_count,
                "skipped": skipped_count,
                "correct": correct_count,
                "incorrect": incorrect_count,
                "accuracy": accuracy,
                "current_streak": user.get("current_streak", 0),
                "best_streak": user.get("best_streak", 0),
                "total_played_all_time": user.get("total_played", 0),
                "total_correct_all_time": user.get("correct_answers", 0),
                "total_skipped_all_time": user.get("skipped", 0),
            })
        
        # Sort by played count (descending)
        user_stats.sort(key=lambda x: x["played"], reverse=True)
        
        # Calculate totals
        total_played = sum(u["played"] for u in user_stats)
        total_skipped = sum(u["skipped"] for u in user_stats)
        total_correct = sum(u["correct"] for u in user_stats)
        active_users = sum(1 for u in user_stats if u["played"] > 0 or u["skipped"] > 0)
        
        return {
            "date": target_date.strftime("%Y-%m-%d"),
            "summary": {
                "total_users": len(all_users),
                "active_users": active_users,
                "total_played": total_played,
                "total_skipped": total_skipped,
                "total_correct": total_correct,
                "overall_accuracy": round((total_correct / total_played * 100), 1) if total_played > 0 else 0
            },
            "user_stats": user_stats
        }
    except HTTPException:
        raise
    except Exception as e:
        logging.error(f"Error getting stats: {e}")
        raise HTTPException(status_code=500, detail=str(e))

# ==================== BULK UPLOAD ENDPOINTS ====================

# Note: UploadFile, File, Form imported at top of file
from fastapi.responses import StreamingResponse
import io
import csv
from openpyxl import Workbook, load_workbook

# Sample data for templates
SAMPLE_DATA = {
    "text_mcq": [
        {"category": "Science", "question_text": "What is H2O commonly known as?", 
         "option_1": "Salt", "option_2": "Water", "option_3": "Sugar", "option_4": "Oil", 
         "correct_answer": "Water", "answer_explanation": "H2O is the chemical formula for water, consisting of 2 hydrogen atoms and 1 oxygen atom.", "difficulty": "easy"},
        {"category": "History", "question_text": "Which civilization built the pyramids?", 
         "option_1": "Roman", "option_2": "Greek", "option_3": "Egyptian", "option_4": "Persian", 
         "correct_answer": "Egyptian", "answer_explanation": "The ancient Egyptians built the pyramids around 2500 BCE as tombs for their pharaohs.", "difficulty": "easy"},
    ],
    "text_input": [
        {"category": "Geography", "question_text": "What is the capital of France?", 
         "correct_answer": "Paris", "alternate_answers": "paris, Paree, paree", "answer_explanation": "Paris has been the capital of France since 987 AD and is known as the 'City of Light'.", "difficulty": "easy"},
        {"category": "Literature", "question_text": "Who wrote 'Hamlet'?", 
         "correct_answer": "Shakespeare", "alternate_answers": "William Shakespeare, shakespear, Shakespear, W. Shakespeare", "answer_explanation": "William Shakespeare wrote Hamlet around 1600. It's considered one of the greatest plays ever written.", "difficulty": "medium"},
    ],
    "image_mcq": [
        {"category": "Art", "image_url": "https://example.com/image1.jpg", 
         "question_text": "Who painted this artwork?", 
         "option_1": "Van Gogh", "option_2": "Picasso", "option_3": "Da Vinci", "option_4": "Monet", 
         "correct_answer": "Da Vinci", "answer_explanation": "Leonardo da Vinci was an Italian Renaissance polymath known for masterpieces like the Mona Lisa.", "difficulty": "medium"},
    ],
    "image_text_input": [
        {"category": "Geography", "image_url": "https://example.com/landmark.jpg", 
         "question_text": "Name this famous landmark", 
         "correct_answer": "Eiffel Tower", "alternate_answers": "eiffel tower, The Eiffel Tower, Tour Eiffel, Eiffel", "answer_explanation": "The Eiffel Tower was built in 1889 for the World's Fair and stands 330 meters tall in Paris.", "difficulty": "easy"},
    ],
    "video_mcq": [
        {"category": "Science", "video_url": "https://example.com/video.mp4", 
         "question_text": "What principle is demonstrated in this video?", 
         "option_1": "Gravity", "option_2": "Magnetism", "option_3": "Electricity", "option_4": "Sound", 
         "correct_answer": "Gravity", "answer_explanation": "Gravity is the force that attracts objects with mass toward each other, as demonstrated by falling objects.", "difficulty": "medium"},
    ],
    "video_text_input": [
        {"category": "Music", "video_url": "https://example.com/music.mp4", 
         "question_text": "What instrument is being played?", 
         "correct_answer": "Piano", "alternate_answers": "piano, Grand Piano, grand piano, Keyboard", "answer_explanation": "The piano was invented around 1700 and is known for its wide range and expressive capabilities.", "difficulty": "easy"},
    ],
}

def get_template_columns(format_type: str) -> List[str]:
    """Get column headers for each format type"""
    base_cols = ["category", "difficulty"]
    
    if format_type == "text_mcq":
        return base_cols + ["question_text", "option_1", "option_2", "option_3", "option_4", "correct_answer", "answer_explanation"]
    elif format_type == "text_input":
        return base_cols + ["question_text", "correct_answer", "alternate_answers", "answer_explanation"]
    elif format_type == "image_mcq":
        return base_cols + ["image_url", "question_text", "option_1", "option_2", "option_3", "option_4", "correct_answer", "answer_explanation"]
    elif format_type == "image_text_input":
        return base_cols + ["image_url", "question_text", "correct_answer", "alternate_answers", "answer_explanation"]
    elif format_type == "video_mcq":
        return base_cols + ["video_url", "question_text", "option_1", "option_2", "option_3", "option_4", "correct_answer", "answer_explanation"]
    elif format_type == "video_text_input":
        return base_cols + ["video_url", "question_text", "correct_answer", "alternate_answers", "answer_explanation"]
    else:
        return base_cols + ["question_text", "correct_answer", "alternate_answers", "answer_explanation"]

@api_router.get("/admin/template/{format_type}")
async def download_template(format_type: str, file_format: str = "xlsx", _: bool = Depends(verify_admin_token)):
    """Download sample template for bulk upload"""
    
    valid_formats = ["text_mcq", "text_input", "image_mcq", "image_text_input", "video_mcq", "video_text_input"]
    if format_type not in valid_formats:
        raise HTTPException(status_code=400, detail=f"Invalid format type. Valid types: {valid_formats}")
    
    columns = get_template_columns(format_type)
    sample_data = SAMPLE_DATA.get(format_type, [])
    
    if file_format == "csv":
        # Generate CSV
        output = io.StringIO()
        writer = csv.DictWriter(output, fieldnames=columns)
        writer.writeheader()
        for row in sample_data:
            writer.writerow({col: row.get(col, "") for col in columns})
        
        output.seek(0)
        return StreamingResponse(
            iter([output.getvalue()]),
            media_type="text/csv",
            headers={"Content-Disposition": f"attachment; filename=template_{format_type}.csv"}
        )
    else:
        # Generate Excel
        wb = Workbook()
        ws = wb.active
        ws.title = format_type
        
        # Add headers
        for col_idx, col_name in enumerate(columns, 1):
            cell = ws.cell(row=1, column=col_idx, value=col_name)
            cell.font = cell.font.copy(bold=True)
        
        # Add sample data
        for row_idx, row_data in enumerate(sample_data, 2):
            for col_idx, col_name in enumerate(columns, 1):
                ws.cell(row=row_idx, column=col_idx, value=row_data.get(col_name, ""))
        
        # Add instructions sheet
        ws_instructions = wb.create_sheet("Instructions")
        instructions = [
            "BULK UPLOAD INSTRUCTIONS",
            "",
            f"Format Type: {format_type.upper()}",
            "",
            "COLUMN DESCRIPTIONS:",
            "- category: The category/topic of the question (e.g., Science, History)",
            "- difficulty: easy, medium, or hard",
            "- question_text: The actual question to display",
            "- correct_answer: The correct answer (must match one of the options for MCQ)",
        ]
        
        if "mcq" in format_type:
            instructions.extend([
                "- option_1 to option_4: The four multiple choice options",
                "",
                "IMPORTANT: correct_answer MUST exactly match one of the options!"
            ])
        
        if "image" in format_type:
            instructions.append("- image_url: Public URL to the image (must be accessible)")
        if "video" in format_type:
            instructions.append("- video_url: Public URL to the video file (mp4 recommended)")
        
        for row_idx, line in enumerate(instructions, 1):
            ws_instructions.cell(row=row_idx, column=1, value=line)
        
        # Save to bytes
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename=template_{format_type}.xlsx"}
        )

@api_router.post("/admin/bulk-upload")
async def bulk_upload_playables(
    file: UploadFile = File(...),
    format_type: str = Form("text_mcq"),
    _: bool = Depends(verify_admin_token)
):
    """Bulk upload playables from Excel or CSV file"""
    
    logging.info(f"Bulk upload called with format_type: {format_type}, filename: {file.filename}")
    
    valid_formats = ["text_mcq", "text_input", "image_mcq", "image_text_input", "video_mcq", "video_text_input"]
    if format_type not in valid_formats:
        raise HTTPException(status_code=400, detail=f"Invalid format type. Valid types: {valid_formats}")
    
    # Read file
    content = await file.read()
    
    try:
        rows = []
        
        if file.filename.endswith('.csv'):
            # Parse CSV
            text_content = content.decode('utf-8')
            reader = csv.DictReader(io.StringIO(text_content))
            rows = list(reader)
        elif file.filename.endswith(('.xlsx', '.xls')):
            # Parse Excel
            wb = load_workbook(io.BytesIO(content))
            ws = wb.active
            
            # Get headers from first row
            headers = [cell.value for cell in ws[1] if cell.value]
            
            # Get data rows
            for row in ws.iter_rows(min_row=2, values_only=True):
                if any(row):  # Skip empty rows
                    row_dict = {headers[i]: row[i] for i in range(len(headers)) if i < len(row)}
                    rows.append(row_dict)
        else:
            raise HTTPException(status_code=400, detail="Invalid file format. Please upload .csv or .xlsx file")
        
        if not rows:
            raise HTTPException(status_code=400, detail="No data found in file")
        
        # Process rows and create playables
        created = []
        errors = []
        
        for idx, row in enumerate(rows, 1):
            try:
                # Helper function to safely get and strip values (handles None from Excel)
                def safe_get(key: str, default: str = '') -> str:
                    value = row.get(key)
                    if value is None:
                        return default
                    return str(value).strip()
                
                # Validate required fields (title removed - no longer required)
                category = safe_get('category')
                question_text = safe_get('question_text')
                correct_answer = safe_get('correct_answer')
                difficulty = safe_get('difficulty', 'medium').lower()
                
                if not category or not correct_answer:
                    errors.append(f"Row {idx}: Missing required fields (category or correct_answer)")
                    continue
                
                # Validate category exists
                category_doc = await db.categories.find_one({"name": {"$regex": f"^{category}$", "$options": "i"}})
                if not category_doc:
                    errors.append(f"Row {idx}: Category '{category}' does not exist. Add it in the Categories tab first.")
                    continue
                
                # Use the exact category name from database
                validated_category = category_doc["name"]
                
                # Build question object
                question = {}
                if question_text:
                    question["text"] = question_text
                
                # Handle image/video URLs
                if "image" in format_type:
                    image_url = safe_get('image_url')
                    if image_url:
                        question["image_base64"] = image_url  # Using same field for compatibility
                    else:
                        errors.append(f"Row {idx}: Image URL required for image format")
                        continue
                
                if "video" in format_type:
                    video_url = safe_get('video_url')
                    if video_url:
                        question["video_url"] = video_url
                    else:
                        errors.append(f"Row {idx}: Video URL required for video format")
                        continue
                
                # Determine type and answer_type using the config
                # Map bulk format_type to playable type
                format_to_type = {
                    "text_mcq": ("text", "mcq"),
                    "text_input": ("text", "text_input"),
                    "image_mcq": ("image_text", "mcq"),
                    "image_text_input": ("image_text", "text_input"),
                    "video_mcq": ("video_text", "mcq"),
                    "video_text_input": ("video_text", "text_input"),
                }
                
                playable_type, requested_answer_type = format_to_type.get(format_type, ("text", "mcq"))
                
                # Validate answer_type using config
                try:
                    answer_type = get_valid_answer_type(playable_type, requested_answer_type)
                except ValueError as e:
                    errors.append(f"Row {idx}: {str(e)}")
                    continue
                
                # Handle MCQ options
                options = None
                if "mcq" in format_type:
                    options = [
                        safe_get('option_1'),
                        safe_get('option_2'),
                        safe_get('option_3'),
                        safe_get('option_4'),
                    ]
                    options = [o for o in options if o]  # Remove empty options
                    
                    if len(options) < 2:
                        errors.append(f"Row {idx}: MCQ requires at least 2 options")
                        continue
                    
                    if correct_answer not in options:
                        errors.append(f"Row {idx}: Correct answer '{correct_answer}' not in options")
                        continue
                
                # Create playable document
                playable_id = f"play_{uuid.uuid4().hex[:12]}"
                answer_explanation = safe_get('answer_explanation') or None
                
                # Parse alternate answers (comma-separated string to list)
                alternate_answers_str = safe_get('alternate_answers')
                alternate_answers = None
                if alternate_answers_str and answer_type == "text_input":
                    alternate_answers = [a.strip() for a in alternate_answers_str.split(',') if a.strip()]
                
                playable_doc = {
                    "playable_id": playable_id,
                    "type": playable_type,
                    "answer_type": answer_type,
                    "category": validated_category,  # Use validated category name
                    "question": question,
                    "options": options,
                    "correct_answer": correct_answer,
                    "alternate_answers": alternate_answers,
                    "answer_explanation": answer_explanation,
                    "difficulty": difficulty if difficulty in ["easy", "medium", "hard"] else "medium",
                    "weight": 0,  # Default weight for bulk uploads
                    "status": "active",  # Default status for new playables
                    "created_at": datetime.now(timezone.utc)
                }
                
                await db.playables.insert_one(playable_doc)
                created.append({"row": idx, "playable_id": playable_id})
                
            except Exception as e:
                errors.append(f"Row {idx}: {str(e)}")
        
        return {
            "success": True,
            "total_rows": len(rows),
            "created_count": len(created),
            "error_count": len(errors),
            "created": created,
            "errors": errors
        }
        
    except Exception as e:
        logging.error(f"Bulk upload error: {e}")
        raise HTTPException(status_code=500, detail=f"Error processing file: {str(e)}")

@api_router.get("/admin/template-formats")
async def get_template_formats(_: bool = Depends(verify_admin_token)):
    """Get list of available template formats with descriptions"""
    return {
        "formats": [
            {"id": "text", "name": "Text", "description": "Text-only question"},
            {"id": "image_text", "name": "Image + Text", "description": "Image with text question"},
            {"id": "video_text", "name": "Video + Text", "description": "Video with text question"},
            {"id": "guess_the_x", "name": "Guess the X", "description": "5 hints • Next hint revealed on wrong answer"},
            {"id": "chess_mate_in_2", "name": "Chess Mate in 2", "description": "Chess puzzle - find mate in 2 moves"},
            {"id": "this_or_that", "name": "This or That", "description": "Two images • Tap to select the correct one"},
        ],
        "answer_types": [
            {"id": "mcq", "name": "Multiple Choice", "description": "4 options to choose from"},
            {"id": "text_input", "name": "Text Input", "description": "User types the answer"},
            {"id": "tap_select", "name": "Tap Select", "description": "User taps to select (for This or That)"},
        ]
    }

# Public endpoint to get playable types (no auth required - for agents)
@api_router.get("/playable-types")
async def get_playable_types():
    """Get all playable types with descriptions (public endpoint for agents)"""
    return {
        "types": [
            {
                "id": "text",
                "name": "Text",
                "description": "Text-only question",
                "supported_answer_types": ["mcq", "text_input"]
            },
            {
                "id": "image_text",
                "name": "Image + Text",
                "description": "Image with text question",
                "supported_answer_types": ["mcq", "text_input"]
            },
            {
                "id": "video_text",
                "name": "Video + Text",
                "description": "Video with text question",
                "valid_answer_types": ["mcq", "text_input"],
                "default_answer_type": "mcq"
            },
            {
                "id": "guess_the_x",
                "name": "Guess the X",
                "description": "5 hints • Next hint revealed on wrong answer",
                "valid_answer_types": ["progressive_reveal"],
                "default_answer_type": "progressive_reveal",
                "answer_type_auto": True,
                "required_fields": ["hints"],
                "hints_count": "3-5"
            },
            {
                "id": "chess_mate_in_2",
                "name": "Chess Mate in 2",
                "description": "Chess puzzle - find mate in 2 moves",
                "valid_answer_types": ["chess_moves"],
                "default_answer_type": "chess_moves",
                "answer_type_auto": True,
                "required_fields": ["fen", "solution"]
            },
            {
                "id": "this_or_that",
                "name": "This or That",
                "description": "Two images • Tap to select the correct one",
                "valid_answer_types": ["tap_select"],
                "default_answer_type": "tap_select",
                "answer_type_auto": True,
                "required_fields": ["image_left_url", "image_right_url", "label_left", "label_right"]
            },
            {
                "id": "wordle",
                "name": "Wordle",
                "description": "5-letter word guessing game",
                "valid_answer_types": ["wordle_grid"],
                "default_answer_type": "wordle_grid",
                "answer_type_auto": True,
                "required_fields": ["correct_answer (5 letters)"]
            }
        ],
        "type_config": {
            "description": "Defines valid answer_types for each playable type. Types with answer_type_auto=True have fixed answer_type.",
            "mapping": {k: {"valid_answer_types": v["valid_answer_types"], "default": v["default_answer_type"]} for k, v in PLAYABLE_TYPE_CONFIG.items()}
        }
    }

# ==================== API DOCUMENTATION ENDPOINT ====================
# ╔═══════════════════════════════════════════════════════════════════════════╗
# ║  IMPORTANT MAINTENANCE INSTRUCTION                                         ║
# ║  ─────────────────────────────────────────────────────────────────────────║
# ║  When adding NEW playable types/formats or modifying the API:              ║
# ║                                                                            ║
# ║  1. Update the "type" enum in playable_schema.required_fields              ║
# ║  2. Add new fields to optional_fields if applicable                        ║
# ║  3. Add example payload in example_payloads section                        ║
# ║  4. Update /app/API_TEMPLATE.md documentation                              ║
# ║  5. Increment the version number                                           ║
# ║                                                                            ║
# ║  This schema is consumed by external agents for API integration.           ║
# ║  Keeping it updated prevents integration errors.                           ║
# ╚═══════════════════════════════════════════════════════════════════════════╝

@api_router.get("/docs/schema")
async def get_api_schema():
    """
    Get API schema documentation for agents/programmatic consumption.
    
    MAINTENANCE: When adding new playable types or API changes:
    1. Update the enum values in this schema
    2. Add example payloads for new types
    3. Update /app/API_TEMPLATE.md
    4. Increment version number
    """
    return {
        "version": "1.3",
        "base_endpoints": {
            "admin_login": "POST /api/admin/login",
            "playables": {
                "list": {
                    "endpoint": "GET /api/admin/playables",
                    "query_params": {
                        "page": {"type": "integer", "default": 1, "description": "Page number (1-indexed)"},
                        "limit": {"type": "integer", "default": 100, "min": 1, "max": 500, "description": "Items per page"},
                        "category": {"type": "string", "optional": True, "description": "Filter by category name"},
                        "type": {"type": "string", "optional": True, "description": "Filter by playable type"},
                        "status": {"type": "string", "default": "active", "enum": ["active", "inactive", "all"], "description": "Filter by status"}
                    },
                    "response": {
                        "playables": "array of playable objects",
                        "count": "number of items in current page",
                        "total": "total number of matching playables",
                        "page": "current page number",
                        "limit": "items per page",
                        "total_pages": "total number of pages",
                        "status_filter": "current status filter applied"
                    }
                },
                "create": "POST /api/admin/add-playable",
                "update": "PUT /api/admin/playables/{playable_id}",
                "delete": "DELETE /api/admin/playables/{playable_id}"
            },
            "categories": {
                "list": "GET /api/admin/categories",
                "create": "POST /api/admin/categories",
                "update": "PATCH /api/admin/categories/{category_id}",
                "delete": "DELETE /api/admin/categories/{category_id}",
                "init": "POST /api/admin/categories/init",
                "valid_icons": "GET /api/admin/valid-icons"
            },
            "bulk_upload": {
                "endpoint": "POST /api/admin/bulk-upload",
                "description": "Upload Excel file with multiple playables",
                "form_data": {
                    "file": "Excel file (.xlsx)",
                    "format_type": "text_mcq | text_input | image_mcq | image_input | video_mcq | video_input"
                }
            },
            "stats": {
                "endpoint": "GET /api/admin/stats",
                "query_params": {
                    "date": {"type": "string", "format": "YYYY-MM-DD", "description": "Date for stats"}
                }
            },
            "reset_user": {
                "endpoint": "POST /api/admin/reset-user-progress",
                "body": {"email": "string"}
            }
        },
        "filter_options": {
            "playable_types": [
                {"value": "text", "label": "Text"},
                {"value": "image", "label": "Image"},
                {"value": "video", "label": "Video"},
                {"value": "image_text", "label": "Image + Text"},
                {"value": "video_text", "label": "Video + Text"},
                {"value": "guess_the_x", "label": "Guess the X"},
                {"value": "chess_mate_in_2", "label": "Chess Puzzle"},
                {"value": "this_or_that", "label": "This or That"},
                {"value": "wordle", "label": "Wordle"}
            ],
            "categories_endpoint": "GET /api/admin/categories (returns list with name, icon, color, playable_count)"
        },
        "database_indexes": {
            "playables": ["playable_id (unique)", "category", "type", "weight", "created_at", "status", "(category, type) compound"],
            "users": ["user_id (unique)", "email (unique)"],
            "user_progress": ["(user_id, playable_id) compound unique"]
        },
        "type_answer_type_config": {
            "description": "Defines valid answer_types for each playable type. Types with single valid_answer_type are auto-determined.",
            "mapping": {k: {"valid_answer_types": v["valid_answer_types"], "default": v["default_answer_type"]} for k, v in PLAYABLE_TYPE_CONFIG.items()}
        },
        "playable_schema": {
            "required_fields": {
                "type": {
                    "type": "string",
                    "enum": ["text", "image", "video", "image_text", "video_text", "guess_the_x", "chess_mate_in_2", "this_or_that"],
                    "description": "Type of playable content",
                    "type_descriptions": {
                        "text": "Text-only question",
                        "image": "Image-based question",
                        "video": "Video-based question",
                        "image_text": "Image with text question",
                        "video_text": "Video with text question",
                        "guess_the_x": "5 hints • Next hint revealed on wrong answer",
                        "chess_mate_in_2": "Chess puzzle - find mate in 2 moves",
                        "this_or_that": "Two images • Tap to select the correct one"
                    }
                },
                "answer_type": {
                    "type": "string",
                    "enum": ["mcq", "text_input", "tap_select", "progressive_reveal", "chess_moves", "wordle_grid"],
                    "description": "How user answers the question. Note: Some types have auto-determined answer_type",
                    "auto_determined_types": {
                        "guess_the_x": "progressive_reveal",
                        "chess_mate_in_2": "chess_moves", 
                        "this_or_that": "tap_select",
                        "wordle": "wordle_grid"
                    }
                },
                "category": {
                    "type": "string",
                    "description": "Must match an existing category name (case-sensitive)"
                },
                "title": {
                    "type": "string",
                    "description": "Display title for the playable"
                },
                "correct_answer": {
                    "type": "string",
                    "description": "The correct answer (for this_or_that: must match label_left or label_right)"
                }
            },
            "optional_fields": {
                "question_text": {
                    "type": "string",
                    "description": "The question text (NOT nested under 'question')"
                },
                "video_url": {
                    "type": "string",
                    "description": "URL to video - MP4 format only (NOT nested under 'question')"
                },
                "video_start": {
                    "type": "integer",
                    "description": "Start time in seconds for video clips"
                },
                "video_end": {
                    "type": "integer",
                    "description": "End time in seconds for video clips"
                },
                "image_url": {
                    "type": "string",
                    "description": "URL to image or base64 data URL (NOT nested under 'question')"
                },
                "image_left_url": {
                    "type": "string",
                    "description": "Left image URL (for this_or_that only)"
                },
                "image_right_url": {
                    "type": "string",
                    "description": "Right image URL (for this_or_that only)"
                },
                "label_left": {
                    "type": "string",
                    "description": "Label for left image - used for answer matching (for this_or_that only)"
                },
                "label_right": {
                    "type": "string",
                    "description": "Label for right image - used for answer matching (for this_or_that only)"
                },
                "options": {
                    "type": "array",
                    "items": "string",
                    "description": "4 options for MCQ (required when answer_type is 'mcq')"
                },
                "alternate_answers": {
                    "type": "array",
                    "items": "string",
                    "description": "Alternative accepted answers for text_input"
                },
                "answer_explanation": {
                    "type": "string",
                    "description": "Explanation shown after answering"
                },
                "hints": {
                    "type": "array",
                    "items": "string",
                    "description": "3-5 progressive hints (for guess_the_x only)"
                },
                "fen": {
                    "type": "string",
                    "description": "Chess position in FEN notation (for chess_mate_in_2 only)"
                },
                "solution": {
                    "type": "array",
                    "items": "string",
                    "description": "Chess moves in UCI format (for chess_mate_in_2 only)"
                },
                "difficulty": {
                    "type": "string",
                    "enum": ["easy", "medium", "hard"],
                    "default": "medium"
                }
            }
        },
        "category_schema": {
            "create": {
                "name": {"type": "string", "required": True},
                "icon": {"type": "string", "required": False, "default": "help-circle", "description": "Valid Ionicons name"},
                "color": {"type": "string", "required": False, "default": "#00FF87", "description": "Hex color (#RGB or #RRGGBB)"}
            },
            "update": {
                "icon": {"type": "string", "required": False},
                "color": {"type": "string", "required": False}
            }
        },
        "important_notes": [
            "Use FLAT fields (question_text, video_url) NOT nested objects (question: {text, video_url})",
            "Category names are case-sensitive and must exist before creating playables",
            "Icons must be valid Ionicons names - check /api/admin/valid-icons",
            "All admin endpoints require Authorization: Bearer <admin_token> header"
        ],
        "example_payloads": {
            "text_mcq": {
                "type": "text",
                "answer_type": "mcq",
                "category": "SCIENCE",
                "title": "Chemistry Basics",
                "question_text": "What is the chemical symbol for Gold?",
                "options": ["Au", "Ag", "Fe", "Cu"],
                "correct_answer": "Au",
                "difficulty": "easy"
            },
            "video_mcq": {
                "type": "video",
                "answer_type": "mcq",
                "category": "MATHS",
                "title": "Math Puzzle",
                "question_text": "Solve this!",
                "video_url": "https://example.com/video.mp4",
                "options": ["7", "14", "10", "24"],
                "correct_answer": "7",
                "difficulty": "medium"
            },
            "guess_the_x": {
                "type": "guess_the_x",
                "answer_type": "text_input",
                "category": "MOVIES",
                "title": "Guess the Movie",
                "question_text": "Guess from these hints",
                "hints": ["Hint 1", "Hint 2", "Hint 3", "Hint 4", "Hint 5"],
                "correct_answer": "The Answer",
                "alternate_answers": ["answer", "the answer"],
                "difficulty": "medium"
            },
            "this_or_that": {
                "type": "this_or_that",
                "answer_type": "tap_select",
                "category": "GENERAL",
                "title": "Logo Recognition",
                "question_text": "Which is the Apple logo?",
                "image_left_url": "https://example.com/apple-logo.png",
                "image_right_url": "https://example.com/samsung-logo.png",
                "label_left": "Apple",
                "label_right": "Samsung",
                "correct_answer": "Apple",
                "difficulty": "easy"
            }
        }
    }

# Include the router in the main app
app.include_router(api_router)

app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=["*"],
    allow_methods=["*"],
    allow_headers=["*"],
)

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

# Note: Startup and shutdown events are defined at the top of the file
# in the TaskManager section. The following duplicate events have been removed:
# - startup_db_client() - indexes are now created in startup_event()
# - shutdown_db_client() - client closure is now handled in shutdown_event()
